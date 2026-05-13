"""
Backend MAGAZYN v5.0 - KOMPLETNA WERSJA
Stack: FastAPI + SQLAlchemy + asyncpg + PostgreSQL

Nowości w v5:
- Migracje bazy (ALTER TABLE) - aplikacja sama dodaje brakujące kolumny
- Kontenery: order_number zamiast supplier; auto-fill ceny z Subiekta
- Endpointy: /api/products/import, /api/products/export, /api/anomalies
- Endpoint: /api/containers/{id}/order-pdf-data, /api/auto-suggest
- Załączniki: /api/containers/{id}/attachments (mock - tylko metadane)
- Lista zakupów grupowana per producent: /api/shopping-list
- Wykres wartości magazynu: /api/stock-value-history
"""

from datetime import date, timedelta, datetime
from typing import List, Optional, Literal, Any
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, Depends, Query, UploadFile, File, Request, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from pydantic import BaseModel, Field, EmailStr
from pydantic_settings import BaseSettings
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from passlib.context import CryptContext
from jose import jwt, JWTError

import io
import csv
import re
import secrets


class Settings(BaseSettings):
    DATABASE_URL: str = ""
    DB_HOST: str = ""
    DB_PORT: int = 5432
    DB_NAME: str = ""
    DB_USER: str = ""
    DB_PASSWORD: str = ""
    
    ALLOWED_ORIGINS: str = "http://localhost:3000"
    
    TABLE_PRODUCTS: str = "subiekt_towary"
    COL_PRODUCT_SKU: str = "symbol"
    COL_PRODUCT_NAME: str = "nazwa"
    COL_PRODUCT_STOCK: str = "stan_dostepny"
    COL_PRODUCT_PRICE: str = "cena_zakupu_netto"
    
    TABLE_ORDERS: str = "sellasist_orders"
    COL_ORDER_ID: str = "order_id"
    COL_ORDER_DATE: str = "order_date"
    COL_ORDER_STATUS: str = "status_name"
    
    TABLE_ORDER_ITEMS: str = "sellasist_order_items"
    COL_ITEM_ORDER_ID: str = "order_id"
    COL_ITEM_SKU: str = "symbol"
    COL_ITEM_QTY: str = "quantity"
    COL_ITEM_EAN: str = "ean"
    
    EXCLUDED_ORDER_STATUSES: str = ""
    
    TABLE_LEAD_TIMES: str = "app_lead_times"
    TABLE_PRODUCT_ATTRS: str = "app_product_attrs"
    TABLE_MANUFACTURERS: str = "app_manufacturers"
    TABLE_CONTAINER_TYPES: str = "app_container_types"
    TABLE_CONTAINERS: str = "app_containers"
    TABLE_CONTAINER_ITEMS: str = "app_container_items"
    TABLE_ATTACHMENTS: str = "app_container_attachments"
    TABLE_USERS: str = "app_users"
    TABLE_AUDIT_LOG: str = "app_audit_log"
    
    DEFAULT_LEAD_TIME_DAYS: int = 90
    
    # Auth - WAŻNE: w produkcji ustaw silny SECRET_KEY w zmiennych środowiskowych Railway!
    SECRET_KEY: str = ""
    JWT_ALGORITHM: str = "HS256"
    JWT_EXPIRE_DAYS: int = 7
    
    # Domyślny admin - tworzy się automatycznie przy pierwszym uruchomieniu
    ADMIN_EMAIL: str = ""
    ADMIN_PASSWORD: str = ""
    # Super-admin - tylko ten email widzi audit log (ustaw ten sam co ADMIN_EMAIL)
    SUPER_ADMIN_EMAIL: str = ""
    
    class Config:
        env_file = ".env"


settings = Settings()


if not settings.DATABASE_URL and settings.DB_HOST:
    from urllib.parse import quote_plus
    pw = quote_plus(settings.DB_PASSWORD)
    user = quote_plus(settings.DB_USER)
    settings.DATABASE_URL = f"postgresql+asyncpg://{user}:{pw}@{settings.DB_HOST}:{settings.DB_PORT}/{settings.DB_NAME}"

if not settings.DATABASE_URL:
    raise RuntimeError("Brak konfiguracji bazy. Ustaw DB_HOST/DB_PORT/DB_NAME/DB_USER/DB_PASSWORD w .env")

# Auth setup
if not settings.SECRET_KEY:
    # Generuje losowy klucz na czas pracy procesu (BEZPIECZNE LOKALNIE, ale w produkcji ustaw stały w env!)
    settings.SECRET_KEY = secrets.token_urlsafe(48)
    print("[WARNING] SECRET_KEY nie ustawiony w env - wygenerowano tymczasowy. W produkcji USTAW go w zmiennych środowiskowych Railway!")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def validate_password_strength(password: str) -> Optional[str]:
    """Zwraca komunikat o błędzie lub None jeśli OK."""
    if len(password) < 8:
        return "Hasło musi mieć minimum 8 znaków"
    if not re.search(r"[A-Z]", password):
        return "Hasło musi zawierać przynajmniej jedną wielką literę"
    if not re.search(r"[0-9]", password):
        return "Hasło musi zawierać przynajmniej jedną cyfrę"
    return None


def create_jwt_token(user_id: int, email: str, role: str) -> str:
    expire = datetime.utcnow() + timedelta(days=settings.JWT_EXPIRE_DAYS)
    payload = {"sub": str(user_id), "email": email, "role": role, "exp": expire}
    return jwt.encode(payload, settings.SECRET_KEY, algorithm=settings.JWT_ALGORITHM)


def decode_jwt_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.JWT_ALGORITHM])
        return payload
    except JWTError:
        return None


def _excluded_status_clause(alias: str = "o") -> str:
    if not settings.EXCLUDED_ORDER_STATUSES.strip():
        return ""
    statuses = [s.strip() for s in settings.EXCLUDED_ORDER_STATUSES.split(",") if s.strip()]
    quoted = ",".join(f"'{s}'" for s in statuses)
    return f"AND {alias}.{settings.COL_ORDER_STATUS} NOT IN ({quoted})"


EXCLUDED_STATUS_FILTER = _excluded_status_clause("o")


engine = create_async_engine(
    settings.DATABASE_URL,
    echo=False,
    pool_pre_ping=True,
    pool_size=2,           # MAX 2 jednoczesne połączenia (home.pl ma limit)
    max_overflow=1,        # +1 awaryjnie = max 3 jednocześnie
    pool_timeout=30,       # czekaj max 30s na wolne połączenie
    pool_recycle=300,      # zamknij połączenia starsze niż 5 min
    connect_args={"timeout": 30, "command_timeout": 30},
)
SessionLocal = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)


async def get_db():
    async with SessionLocal() as session:
        yield session


async def add_column_if_missing(conn, table: str, column: str, definition: str):
    """Dodaje kolumnę jeśli nie istnieje. Pomija jak istnieje."""
    try:
        await conn.execute(text(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {definition}"))
    except Exception as e:
        print(f"[migration] {table}.{column}: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    async with engine.begin() as conn:
        # Lead times
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_LEAD_TIMES} (
                sku VARCHAR(255) PRIMARY KEY,
                lead_time_days INTEGER NOT NULL DEFAULT {settings.DEFAULT_LEAD_TIME_DAYS},
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        
        # Producenci
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_MANUFACTURERS} (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL UNIQUE,
                color VARCHAR(20) DEFAULT '#6b7280',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        await add_column_if_missing(conn, settings.TABLE_MANUFACTURERS, "email", "VARCHAR(255)")
        
        # Typy kontenerów
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_CONTAINER_TYPES} (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL UNIQUE,
                capacity_cbm DECIMAL(8,2) NOT NULL,
                sort_order INTEGER DEFAULT 0
            )
        """))
        
        await conn.execute(text(f"""
            INSERT INTO {settings.TABLE_CONTAINER_TYPES} (name, capacity_cbm, sort_order) 
            VALUES ('20''', 33.0, 1), ('40''', 67.0, 2), ('40''HQ', 76.0, 3)
            ON CONFLICT (name) DO NOTHING
        """))
        
        # Atrybuty produktów
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_PRODUCT_ATTRS} (
                sku VARCHAR(255) PRIMARY KEY,
                cbm_per_unit DECIMAL(8,4) DEFAULT 0,
                manufacturer_id INTEGER REFERENCES {settings.TABLE_MANUFACTURERS}(id) ON DELETE SET NULL,
                seasonality_enabled BOOLEAN DEFAULT FALSE,
                force_visible BOOLEAN DEFAULT FALSE,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        # Migracja: ulubione
        await add_column_if_missing(conn, settings.TABLE_PRODUCT_ATTRS, "is_favorite", "BOOLEAN DEFAULT FALSE")
        # Migracja: EAN
        await add_column_if_missing(conn, settings.TABLE_PRODUCT_ATTRS, "ean", "VARCHAR(50)")
        # Migracja: ręczne wymuszenie statusu
        await add_column_if_missing(conn, settings.TABLE_PRODUCT_ATTRS, "forced_status", "VARCHAR(30)")
        
        # Kontenery
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_CONTAINERS} (
                id SERIAL PRIMARY KEY,
                container_number VARCHAR(100) NOT NULL,
                container_type_id INTEGER REFERENCES {settings.TABLE_CONTAINER_TYPES}(id),
                manufacturer_id INTEGER REFERENCES {settings.TABLE_MANUFACTURERS}(id),
                supplier VARCHAR(255),
                order_date DATE NOT NULL,
                eta_date DATE NOT NULL,
                status VARCHAR(30) NOT NULL DEFAULT 'ORDERED',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT chk_status CHECK (status IN ('ORDERED','IN_PRODUCTION','IN_TRANSIT','DELIVERED'))
            )
        """))
        # Migracja: order_number
        await add_column_if_missing(conn, settings.TABLE_CONTAINERS, "order_number", "VARCHAR(100)")
        
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_CONTAINER_ITEMS} (
                id SERIAL PRIMARY KEY,
                container_id INTEGER NOT NULL REFERENCES {settings.TABLE_CONTAINERS}(id) ON DELETE CASCADE,
                sku VARCHAR(255) NOT NULL,
                quantity INTEGER NOT NULL CHECK (quantity > 0),
                unit_cost DECIMAL(10,2)
            )
        """))
        
        # Załączniki kontenerów (mock - tylko metadane bo plik storage to inny temat)
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_ATTACHMENTS} (
                id SERIAL PRIMARY KEY,
                container_id INTEGER NOT NULL REFERENCES {settings.TABLE_CONTAINERS}(id) ON DELETE CASCADE,
                filename VARCHAR(255) NOT NULL,
                file_type VARCHAR(50),
                file_size VARCHAR(50),
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        
        # Użytkownicy
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_USERS} (
                id SERIAL PRIMARY KEY,
                email VARCHAR(255) NOT NULL UNIQUE,
                password_hash VARCHAR(255) NOT NULL,
                full_name VARCHAR(255),
                role VARCHAR(20) NOT NULL DEFAULT 'VIEWER',
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                CONSTRAINT chk_role CHECK (role IN ('ADMIN', 'IMPORT', 'VIEWER'))
            )
        """))
        
        # Audit log - kto co kiedy
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {settings.TABLE_AUDIT_LOG} (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES {settings.TABLE_USERS}(id) ON DELETE SET NULL,
                user_email VARCHAR(255),
                action VARCHAR(100) NOT NULL,
                resource_type VARCHAR(50),
                resource_id VARCHAR(255),
                details TEXT,
                ip_address VARCHAR(45),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        await conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_audit_log_user ON {settings.TABLE_AUDIT_LOG}(user_id)"))
        await conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_audit_log_created ON {settings.TABLE_AUDIT_LOG}(created_at DESC)"))
        
        # Tworzenie domyślnego admina jeśli ustawione w env i nie ma żadnego użytkownika
        if settings.ADMIN_EMAIL and settings.ADMIN_PASSWORD:
            count_result = await conn.execute(text(f"SELECT COUNT(*) FROM {settings.TABLE_USERS}"))
            user_count = count_result.scalar()
            if user_count == 0:
                pwd_err = validate_password_strength(settings.ADMIN_PASSWORD)
                if pwd_err:
                    print(f"[ERROR] ADMIN_PASSWORD słabe: {pwd_err}. Admin NIE utworzony.")
                else:
                    hashed = hash_password(settings.ADMIN_PASSWORD)
                    await conn.execute(
                        text(f"INSERT INTO {settings.TABLE_USERS} (email, password_hash, full_name, role) VALUES (:e, :h, :n, 'ADMIN')"),
                        {"e": settings.ADMIN_EMAIL, "h": hashed, "n": "Administrator"}
                    )
                    print(f"[INFO] Utworzono domyślnego admina: {settings.ADMIN_EMAIL}")
        
        # Indeksy
        try:
            await conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_sellasist_items_symbol_lower ON {settings.TABLE_ORDER_ITEMS} (LOWER(TRIM({settings.COL_ITEM_SKU})))"))
            await conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_subiekt_towary_symbol_lower ON {settings.TABLE_PRODUCTS} (LOWER(TRIM({settings.COL_PRODUCT_SKU})))"))
        except Exception as e:
            print(f"[indexes] {e}")
    
    yield
    await engine.dispose()


app = FastAPI(title="Magazyn API v5", version="5.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS.split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def audit_middleware(request: Request, call_next):
    """
    Middleware który automatycznie loguje wszystkie mutacje (POST/PUT/PATCH/DELETE).
    Nie blokuje żądań - logi zapisywane po odpowiedzi.
    """
    response = await call_next(request)
    
    # Loguj tylko mutacje które się udały (2xx)
    if request.method in ("POST", "PUT", "PATCH", "DELETE") and 200 <= response.status_code < 300:
        # Pomiń endpointy auth (logowanie itp. - mają swój log)
        path = request.url.path
        if any(skip in path for skip in ["/auth/login", "/auth/logout", "/auth/me/password"]):
            return response
        
        # Pobierz token żeby zidentyfikować usera
        token_str = None
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token_str = auth_header[7:]
        # Fallback: token w query param (download)
        if not token_str:
            token_str = request.query_params.get("token")
        
        user_id = None
        user_email = None
        if token_str:
            payload = decode_jwt_token(token_str)
            if payload:
                user_id = int(payload.get("sub", 0)) or None
                user_email = payload.get("email")
        
        # Mapuj metodę + ścieżkę na czytelną akcję
        method_labels = {"POST": "CREATED", "PUT": "UPDATED", "PATCH": "UPDATED", "DELETE": "DELETED"}
        action_label = method_labels.get(request.method, request.method)
        
        # Wyciągnij typ zasobu ze ścieżki (/api/containers/5 → containers)
        path_parts = [p for p in path.strip("/").split("/") if p and p != "api"]
        resource_type = path_parts[0] if path_parts else "unknown"
        resource_id = path_parts[1] if len(path_parts) > 1 else None
        
        action = f"{resource_type.upper()}_{action_label}"
        
        try:
            async with SessionLocal() as db:
                await db.execute(
                    text(f"""
                        INSERT INTO {settings.TABLE_AUDIT_LOG} (user_id, user_email, action, resource_type, resource_id, details)
                        VALUES (:uid, :email, :action, :rtype, :rid, :details)
                    """),
                    {
                        "uid": user_id,
                        "email": user_email,
                        "action": action,
                        "rtype": resource_type,
                        "rid": str(resource_id) if resource_id else None,
                        "details": f"{request.method} {path}",
                    }
                )
                await db.commit()
        except Exception as e:
            print(f"[audit_middleware] {e}")
    
    return response


# ========== MODELE ==========
ContainerStatus = Literal["ORDERED", "IN_PRODUCTION", "IN_TRANSIT", "DELIVERED"]
ProductStatus = Literal["ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"]
UserRole = Literal["ADMIN", "IMPORT", "VIEWER"]


# ========== USER MODELS ==========
class UserCreate(BaseModel):
    email: str = Field(..., min_length=3, max_length=255)
    password: str = Field(..., min_length=8)
    full_name: Optional[str] = None
    role: UserRole = "VIEWER"


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role: Optional[UserRole] = None
    is_active: Optional[bool] = None


class PasswordChange(BaseModel):
    current_password: Optional[str] = None  # opcjonalne - admin może zmieniać bez tego
    new_password: str = Field(..., min_length=8)


class UserOut(BaseModel):
    id: int
    email: str
    full_name: Optional[str]
    role: UserRole
    is_active: bool
    created_at: datetime
    last_login: Optional[datetime]


class LoginRequest(BaseModel):
    email: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut


class AuditLogEntry(BaseModel):
    id: int
    user_id: Optional[int]
    user_email: Optional[str]
    action: str
    resource_type: Optional[str]
    resource_id: Optional[str]
    details: Optional[str]
    ip_address: Optional[str]
    created_at: datetime


class IncomingDelivery(BaseModel):
    container_id: int
    container_number: str
    eta_date: date
    quantity: int
    status: ContainerStatus


class ProductSummary(BaseModel):
    sku: str
    name: str
    stock: int
    stock_value: float
    purchase_price: float = 0  # cena zakupu z Subiekta
    stock_in_transit: int
    product_status: ProductStatus
    cbm_per_unit: float
    manufacturer_id: Optional[int]
    manufacturer_name: Optional[str]
    manufacturer_color: Optional[str] = None
    seasonality_enabled: bool
    is_favorite: bool = False
    ean: Optional[str] = None
    forced_status: Optional[str] = None  # gdy ustawione: produkt ma wymuszony status
    lead_time_days: int
    sales_1m: int
    sales_2m: int
    sales_3m: int
    sales_4m: int
    sales_yoy_30d: int
    sales_yoy_next_30d: int
    avg_monthly_weighted: float
    months_of_stock: float
    days_until_empty: int
    days_until_order: int
    empty_date: date
    order_date: date
    status: str
    incoming_deliveries: List[IncomingDelivery] = []


class LeadTimeUpdate(BaseModel):
    lead_time_days: int = Field(..., ge=1, le=365)


class ProductAttrsUpdate(BaseModel):
    cbm_per_unit: Optional[float] = Field(None, ge=0)
    manufacturer_id: Optional[int] = None
    seasonality_enabled: Optional[bool] = None
    ean: Optional[str] = None
    forced_status: Optional[str] = None  # "ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE", lub None (auto)


class ManufacturerIn(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    color: str = "#6b7280"
    notes: Optional[str] = None
    email: Optional[str] = None


class ManufacturerOut(ManufacturerIn):
    id: int


class ContainerTypeIn(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    capacity_cbm: float = Field(..., gt=0)
    sort_order: int = 0


class ContainerTypeOut(ContainerTypeIn):
    id: int


class ContainerItemIn(BaseModel):
    sku: str
    quantity: int = Field(..., gt=0)
    unit_cost: Optional[float] = None


class ContainerCreate(BaseModel):
    container_number: str
    order_number: Optional[str] = None
    container_type_id: Optional[int] = None
    manufacturer_id: Optional[int] = None
    order_date: date
    eta_date: date
    status: ContainerStatus = "ORDERED"
    notes: Optional[str] = None
    items: List[ContainerItemIn] = Field(..., min_length=1)


class ContainerUpdate(BaseModel):
    container_number: Optional[str] = None
    order_number: Optional[str] = None
    container_type_id: Optional[int] = None
    manufacturer_id: Optional[int] = None
    order_date: Optional[date] = None
    eta_date: Optional[date] = None
    status: Optional[ContainerStatus] = None
    notes: Optional[str] = None
    items: Optional[List[ContainerItemIn]] = None


class ContainerItemOut(ContainerItemIn):
    id: int
    product_name: Optional[str] = None
    cbm_per_unit: float = 0
    total_cbm: float = 0


class AttachmentOut(BaseModel):
    id: int
    filename: str
    file_type: Optional[str]
    file_size: Optional[str]
    uploaded_at: datetime


class AttachmentCreate(BaseModel):
    filename: str
    file_type: Optional[str] = None
    file_size: Optional[str] = None


class ContainerOut(BaseModel):
    id: int
    container_number: str
    order_number: Optional[str] = None
    container_type_id: Optional[int]
    container_type_name: Optional[str]
    container_capacity_cbm: Optional[float]
    manufacturer_id: Optional[int]
    manufacturer_name: Optional[str]
    manufacturer_color: Optional[str] = None
    order_date: date
    eta_date: date
    status: ContainerStatus
    notes: Optional[str]
    items: List[ContainerItemOut]
    attachments: List[AttachmentOut] = []
    total_units: int
    total_cbm: float
    fill_percentage: Optional[float]
    total_value: float


class StockProjectionPoint(BaseModel):
    date: date
    stock: int
    event: Optional[str] = None


class Anomaly(BaseModel):
    sku: str
    name: str
    severity: Literal["high", "medium", "low"]
    type: Literal["sales_spike", "sales_drop", "stock_drain"]
    message: str
    sales_1m: int
    sales_3m_avg: float
    change_pct: float


class ImportRow(BaseModel):
    sku: str
    cbm: Optional[float] = None
    manufacturer_name: Optional[str] = None
    lead_time_days: Optional[int] = None
    seasonality_enabled: Optional[bool] = None


class ImportResult(BaseModel):
    total: int
    updated: int
    skipped: int
    errors: List[str] = []


class ShoppingListGroup(BaseModel):
    manufacturer_id: Optional[int]
    manufacturer_name: Optional[str]
    manufacturer_color: Optional[str]
    manufacturer_email: Optional[str]
    products: List[dict]
    total_skus: int


# ========== SQL ==========
SALES_QUERY = f"""
WITH sales_periods AS (
    SELECT
        LOWER(TRIM(oi.{settings.COL_ITEM_SKU})) AS sku_normalized,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '30 days'  THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_1m,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '60 days'  THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_2m,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '90 days'  THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_3m,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '120 days' THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_4m,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '365 days' THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_12m
    FROM {settings.TABLE_ORDER_ITEMS} oi
    JOIN {settings.TABLE_ORDERS} o ON o.{settings.COL_ORDER_ID} = oi.{settings.COL_ITEM_ORDER_ID}
    WHERE o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '365 days'
      {EXCLUDED_STATUS_FILTER}
    GROUP BY LOWER(TRIM(oi.{settings.COL_ITEM_SKU}))
),
sales_yoy AS (
    SELECT
        LOWER(TRIM(oi.{settings.COL_ITEM_SKU})) AS sku_normalized,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '395 days' AND o.{settings.COL_ORDER_DATE} < NOW() - INTERVAL '365 days' THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_yoy_30d,
        SUM(CASE WHEN o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '365 days' AND o.{settings.COL_ORDER_DATE} < NOW() - INTERVAL '335 days' THEN oi.{settings.COL_ITEM_QTY} ELSE 0 END) AS qty_yoy_next_30d
    FROM {settings.TABLE_ORDER_ITEMS} oi
    JOIN {settings.TABLE_ORDERS} o ON o.{settings.COL_ORDER_ID} = oi.{settings.COL_ITEM_ORDER_ID}
    WHERE o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '395 days'
      AND o.{settings.COL_ORDER_DATE} < NOW() - INTERVAL '335 days'
      {EXCLUDED_STATUS_FILTER}
    GROUP BY LOWER(TRIM(oi.{settings.COL_ITEM_SKU}))
)
SELECT
    p.{settings.COL_PRODUCT_SKU} AS sku,
    p.{settings.COL_PRODUCT_NAME} AS name,
    COALESCE(p.{settings.COL_PRODUCT_STOCK}, 0)::int AS stock,
    COALESCE(p.{settings.COL_PRODUCT_PRICE}, 0)::float AS price,
    COALESCE(lt.lead_time_days, :default_lead_time)::int AS lead_time_days,
    COALESCE(pa.cbm_per_unit, 0)::float AS cbm_per_unit,
    pa.manufacturer_id,
    m.name AS manufacturer_name,
    m.color AS manufacturer_color,
    COALESCE(pa.seasonality_enabled, FALSE) AS seasonality_enabled,
    COALESCE(pa.is_favorite, FALSE) AS is_favorite,
    pa.ean AS ean,
    pa.forced_status AS forced_status,
    COALESCE(pa.force_visible, FALSE) AS force_visible,
    COALESCE(sp.qty_1m, 0)::int AS sales_1m_total,
    COALESCE(sp.qty_2m, 0)::int AS sales_2m_total,
    COALESCE(sp.qty_3m, 0)::int AS sales_3m_total,
    COALESCE(sp.qty_4m, 0)::int AS sales_4m_total,
    COALESCE(sp.qty_12m, 0)::int AS sales_12m_total,
    COALESCE(sy.qty_yoy_30d, 0)::int AS sales_yoy_30d,
    COALESCE(sy.qty_yoy_next_30d, 0)::int AS sales_yoy_next_30d
FROM {settings.TABLE_PRODUCTS} p
LEFT JOIN sales_periods sp ON sp.sku_normalized = LOWER(TRIM(p.{settings.COL_PRODUCT_SKU}))
LEFT JOIN sales_yoy sy ON sy.sku_normalized = LOWER(TRIM(p.{settings.COL_PRODUCT_SKU}))
LEFT JOIN {settings.TABLE_LEAD_TIMES} lt ON lt.sku = p.{settings.COL_PRODUCT_SKU}
LEFT JOIN {settings.TABLE_PRODUCT_ATTRS} pa ON pa.sku = p.{settings.COL_PRODUCT_SKU}
LEFT JOIN {settings.TABLE_MANUFACTURERS} m ON m.id = pa.manufacturer_id
ORDER BY p.{settings.COL_PRODUCT_SKU};
"""


INCOMING_QUERY = f"""
SELECT ci.sku, c.id AS container_id, c.container_number, c.eta_date, c.status, ci.quantity
FROM {settings.TABLE_CONTAINER_ITEMS} ci
JOIN {settings.TABLE_CONTAINERS} c ON c.id = ci.container_id
WHERE c.status != 'DELIVERED'
ORDER BY c.eta_date ASC;
"""


# ========== LOGIKA ==========
def classify_product(row: dict) -> str:
    # Ręczne wymuszenie statusu ma najwyższy priorytet
    forced = row.get("forced_status")
    if forced and forced in ("ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"):
        return forced
    
    stock = row["stock"]
    sales_12m = row["sales_12m_total"]
    if row.get("force_visible", False):
        return "ACTIVE"
    if stock > 0 and sales_12m > 0:
        return "ACTIVE"
    if stock == 0 and sales_12m > 0:
        return "ACTIVE_NO_STOCK"
    if stock > 0 and sales_12m == 0:
        return "DEAD_STOCK"
    return "INACTIVE"


def calculate_forecast(row: dict, incoming: List[dict]) -> ProductSummary:
    sales_1m = row["sales_1m_total"]
    sales_2m_avg = row["sales_2m_total"] / 2
    sales_3m_avg = row["sales_3m_total"] / 3
    sales_4m_avg = row["sales_4m_total"] / 4
    
    avg_monthly = sales_1m * 0.4 + sales_2m_avg * 0.3 + sales_3m_avg * 0.2 + sales_4m_avg * 0.1
    base_daily_sales = avg_monthly / 30 if avg_monthly > 0 else 0
    
    today = date.today()
    eta_map = {}
    stock_in_transit = 0
    incoming_deliveries = []
    
    for inc in incoming:
        if inc["eta_date"] >= today:
            eta_map.setdefault(inc["eta_date"], 0)
            eta_map[inc["eta_date"]] += inc["quantity"]
            stock_in_transit += inc["quantity"]
            incoming_deliveries.append(IncomingDelivery(
                container_id=inc["container_id"],
                container_number=inc["container_number"],
                eta_date=inc["eta_date"],
                quantity=inc["quantity"],
                status=inc["status"],
            ))
    
    current_stock = float(row["stock"])
    days_until_empty = 9999
    
    if base_daily_sales > 0:
        for offset in range(0, 730):
            check_date = today + timedelta(days=offset)
            if check_date in eta_map:
                current_stock += eta_map[check_date]
            current_stock -= base_daily_sales
            if current_stock <= 0:
                days_until_empty = offset
                break
    
    empty_date = today + timedelta(days=days_until_empty)
    order_date = empty_date - timedelta(days=row["lead_time_days"])
    days_until_order = (order_date - today).days
    
    if days_until_order <= 0 and days_until_empty < row["lead_time_days"]:
        status = "KRYTYCZNY"
    elif days_until_order <= 7:
        status = "ZAMOW_TERAZ"
    elif days_until_order <= 30:
        status = "ZAMOW_WKROTCE"
    else:
        status = "OK"
    
    total_available = row["stock"] + stock_in_transit
    months_of_stock = (total_available / avg_monthly) if avg_monthly > 0 else 999.0
    price = row.get("price", 0)
    
    return ProductSummary(
        sku=row["sku"],
        name=row["name"] or "",
        stock=row["stock"],
        stock_value=round(row["stock"] * price, 2),
        purchase_price=round(price, 2),
        stock_in_transit=stock_in_transit,
        product_status=classify_product(row),
        cbm_per_unit=row.get("cbm_per_unit", 0),
        manufacturer_id=row.get("manufacturer_id"),
        manufacturer_name=row.get("manufacturer_name"),
        manufacturer_color=row.get("manufacturer_color"),
        seasonality_enabled=row.get("seasonality_enabled", False),
        is_favorite=row.get("is_favorite", False),
        ean=row.get("ean"),
        forced_status=row.get("forced_status"),
        lead_time_days=row["lead_time_days"],
        sales_1m=sales_1m,
        sales_2m=round(sales_2m_avg),
        sales_3m=round(sales_3m_avg),
        sales_4m=round(sales_4m_avg),
        sales_yoy_30d=row.get("sales_yoy_30d", 0),
        sales_yoy_next_30d=row.get("sales_yoy_next_30d", 0),
        avg_monthly_weighted=round(avg_monthly, 1),
        months_of_stock=round(months_of_stock, 1),
        days_until_empty=days_until_empty,
        days_until_order=days_until_order,
        empty_date=empty_date,
        order_date=order_date,
        status=status,
        incoming_deliveries=sorted(incoming_deliveries, key=lambda d: d.eta_date),
    )


async def auto_deliver_containers(db: AsyncSession):
    await db.execute(text(f"""
        UPDATE {settings.TABLE_CONTAINERS}
        SET status = 'DELIVERED', updated_at = CURRENT_TIMESTAMP
        WHERE status = 'IN_TRANSIT' AND eta_date <= CURRENT_DATE
    """))
    await db.commit()


# ============================================================
# AUTH - dependency, guards, endpointy
# ============================================================

class CurrentUser(BaseModel):
    id: int
    email: str
    role: str
    full_name: Optional[str] = None


class UserCreate(BaseModel):
    email: str = Field(..., min_length=5, max_length=255)
    password: str = Field(..., min_length=8)
    full_name: str = Field(..., min_length=1, max_length=255)
    role: Literal["ADMIN", "IMPORT", "VIEWER"] = "VIEWER"


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role: Optional[Literal["ADMIN", "IMPORT", "VIEWER"]] = None
    is_active: Optional[bool] = None


class PasswordChange(BaseModel):
    current_password: str
    new_password: str = Field(..., min_length=8)


class AdminPasswordReset(BaseModel):
    new_password: str = Field(..., min_length=8)


class UserOut(BaseModel):
    id: int
    email: str
    full_name: Optional[str]
    role: str
    is_active: bool
    is_super_admin: bool = False  # tylko ten email widzi audit log
    created_at: datetime
    last_login: Optional[datetime]


class LoginRequest(BaseModel):
    email: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut


class AuditLogOut(BaseModel):
    id: int
    user_id: Optional[int]
    user_email: Optional[str]
    action: str
    resource_type: Optional[str]
    resource_id: Optional[str]
    details: Optional[str]
    created_at: datetime


async def get_current_user(token: Optional[str] = Depends(oauth2_scheme), db: AsyncSession = Depends(get_db)) -> CurrentUser:
    """Wymusza zalogowanego użytkownika - zwraca CurrentUser lub rzuca 401."""
    if not token:
        raise HTTPException(status_code=401, detail="Wymagane logowanie", headers={"WWW-Authenticate": "Bearer"})
    
    payload = decode_jwt_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Token wygasł lub nieprawidłowy", headers={"WWW-Authenticate": "Bearer"})
    
    user_id = int(payload.get("sub", 0))
    if not user_id:
        raise HTTPException(status_code=401, detail="Nieprawidłowy token")
    
    # Sprawdź czy user nadal istnieje i jest aktywny
    r = await db.execute(text(f"SELECT id, email, role, full_name, is_active FROM {settings.TABLE_USERS} WHERE id = :id"), {"id": user_id})
    u = r.first()
    if not u:
        raise HTTPException(status_code=401, detail="Użytkownik nie istnieje")
    if not u.is_active:
        raise HTTPException(status_code=403, detail="Konto deaktywowane")
    
    return CurrentUser(id=u.id, email=u.email, role=u.role, full_name=u.full_name)


def require_role(*allowed_roles):
    """Dependency factory: wymusza określoną rolę."""
    async def checker(user: CurrentUser = Depends(get_current_user)) -> CurrentUser:
        if user.role not in allowed_roles:
            raise HTTPException(403, f"Wymagana rola: {' lub '.join(allowed_roles)}. Twoja rola: {user.role}")
        return user
    return checker


require_admin = require_role("ADMIN")
require_import_or_admin = require_role("ADMIN", "IMPORT")


async def log_audit(db: AsyncSession, user: Optional[CurrentUser], action: str, resource_type: Optional[str] = None, resource_id: Optional[str] = None, details: Optional[str] = None):
    """Zapisuje wpis do audit log. Errory ignorowane (nie blokują głównej operacji)."""
    try:
        await db.execute(
            text(f"""
                INSERT INTO {settings.TABLE_AUDIT_LOG} (user_id, user_email, action, resource_type, resource_id, details)
                VALUES (:uid, :email, :action, :rtype, :rid, :details)
            """),
            {
                "uid": user.id if user else None,
                "email": user.email if user else None,
                "action": action,
                "rtype": resource_type,
                "rid": str(resource_id) if resource_id else None,
                "details": details,
            }
        )
        await db.commit()
    except Exception as e:
        print(f"[audit] {e}")


# ===== ENDPOINTY AUTH =====
@app.post("/api/auth/login", response_model=LoginResponse)
async def login(payload: LoginRequest, db: AsyncSession = Depends(get_db)):
    """Logowanie - zwraca JWT token + dane użytkownika."""
    r = await db.execute(
        text(f"SELECT id, email, password_hash, full_name, role, is_active, created_at, last_login FROM {settings.TABLE_USERS} WHERE LOWER(email) = LOWER(:email)"),
        {"email": payload.email.strip()}
    )
    u = r.first()
    if not u or not verify_password(payload.password, u.password_hash):
        # Logujemy próbę
        await log_audit(db, None, "LOGIN_FAILED", "user", payload.email, "Nieprawidłowy email lub hasło")
        raise HTTPException(401, "Nieprawidłowy email lub hasło")
    
    if not u.is_active:
        await log_audit(db, None, "LOGIN_BLOCKED", "user", payload.email, "Konto deaktywowane")
        raise HTTPException(403, "Konto zostało deaktywowane")
    
    # Update last_login
    await db.execute(
        text(f"UPDATE {settings.TABLE_USERS} SET last_login = CURRENT_TIMESTAMP WHERE id = :id"),
        {"id": u.id}
    )
    await db.commit()
    
    token = create_jwt_token(u.id, u.email, u.role)
    
    user_out = UserOut(
        id=u.id, email=u.email, full_name=u.full_name, role=u.role,
        is_active=u.is_active, created_at=u.created_at, last_login=datetime.now(),
        is_super_admin=bool(settings.SUPER_ADMIN_EMAIL and u.email.lower() == settings.SUPER_ADMIN_EMAIL.strip().lower())
    )
    
    # Audit log
    fake_user = CurrentUser(id=u.id, email=u.email, role=u.role, full_name=u.full_name)
    await log_audit(db, fake_user, "LOGIN", "user", str(u.id))
    
    return LoginResponse(access_token=token, user=user_out)


@app.get("/api/auth/me", response_model=UserOut)
async def get_me(user: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Zwraca dane zalogowanego użytkownika."""
    r = await db.execute(
        text(f"SELECT id, email, full_name, role, is_active, created_at, last_login FROM {settings.TABLE_USERS} WHERE id = :id"),
        {"id": user.id}
    )
    u = r.first()
    if not u:
        raise HTTPException(404, "Użytkownik nie istnieje")
    
    data = dict(u._mapping)
    super_email = settings.SUPER_ADMIN_EMAIL.strip().lower()
    data["is_super_admin"] = bool(super_email and data["email"].lower() == super_email)
    return UserOut(**data)


@app.put("/api/auth/me/password", status_code=204)
async def change_my_password(payload: PasswordChange, user: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Zmiana własnego hasła - wymaga obecnego hasła."""
    err = validate_password_strength(payload.new_password)
    if err:
        raise HTTPException(400, err)
    
    r = await db.execute(text(f"SELECT password_hash FROM {settings.TABLE_USERS} WHERE id = :id"), {"id": user.id})
    u = r.first()
    if not verify_password(payload.current_password, u.password_hash):
        raise HTTPException(400, "Aktualne hasło jest nieprawidłowe")
    
    new_hash = hash_password(payload.new_password)
    await db.execute(text(f"UPDATE {settings.TABLE_USERS} SET password_hash = :h WHERE id = :id"), {"h": new_hash, "id": user.id})
    await db.commit()
    
    await log_audit(db, user, "PASSWORD_CHANGED", "user", str(user.id))


# ===== ENDPOINTY USERS (tylko admin) =====
@app.get("/api/users", response_model=List[UserOut])
async def list_users(admin: CurrentUser = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Lista wszystkich użytkowników - tylko admin."""
    r = await db.execute(text(f"SELECT id, email, full_name, role, is_active, created_at, last_login FROM {settings.TABLE_USERS} ORDER BY created_at DESC"))
    return [UserOut(**dict(row._mapping)) for row in r]


@app.post("/api/users", response_model=UserOut, status_code=201)
async def create_user(payload: UserCreate, admin: CurrentUser = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Tworzy nowego użytkownika - tylko admin."""
    err = validate_password_strength(payload.password)
    if err:
        raise HTTPException(400, err)
    
    # Sprawdź czy email zajęty
    r = await db.execute(text(f"SELECT id FROM {settings.TABLE_USERS} WHERE LOWER(email) = LOWER(:email)"), {"email": payload.email.strip()})
    if r.first():
        raise HTTPException(409, "Użytkownik z tym emailem już istnieje")
    
    pwd_hash = hash_password(payload.password)
    r = await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_USERS} (email, password_hash, full_name, role, is_active)
            VALUES (:e, :h, :n, :r, TRUE) RETURNING id, email, full_name, role, is_active, created_at, last_login
        """),
        {"e": payload.email.strip(), "h": pwd_hash, "n": payload.full_name, "r": payload.role}
    )
    u = r.first()
    await db.commit()
    
    await log_audit(db, admin, "USER_CREATED", "user", str(u.id), f"{payload.email} ({payload.role})")
    return UserOut(**dict(u._mapping))


@app.patch("/api/users/{uid}", response_model=UserOut)
async def update_user(uid: int, payload: UserUpdate, admin: CurrentUser = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Aktualizuje użytkownika - tylko admin."""
    if uid == admin.id and payload.role and payload.role != "ADMIN":
        raise HTTPException(400, "Nie możesz odebrać sobie roli admina!")
    if uid == admin.id and payload.is_active is False:
        raise HTTPException(400, "Nie możesz deaktywować własnego konta!")
    
    updates = []
    params = {"id": uid}
    if payload.full_name is not None:
        updates.append("full_name = :name")
        params["name"] = payload.full_name
    if payload.role is not None:
        updates.append("role = :role")
        params["role"] = payload.role
    if payload.is_active is not None:
        updates.append("is_active = :active")
        params["active"] = payload.is_active
    
    if updates:
        await db.execute(text(f"UPDATE {settings.TABLE_USERS} SET {', '.join(updates)} WHERE id = :id"), params)
        await db.commit()
    
    r = await db.execute(text(f"SELECT id, email, full_name, role, is_active, created_at, last_login FROM {settings.TABLE_USERS} WHERE id = :id"), {"id": uid})
    u = r.first()
    if not u:
        raise HTTPException(404, "Użytkownik nie znaleziony")
    
    await log_audit(db, admin, "USER_UPDATED", "user", str(uid), str(payload.model_dump(exclude_none=True)))
    return UserOut(**dict(u._mapping))


@app.put("/api/users/{uid}/password", status_code=204)
async def reset_user_password(uid: int, payload: AdminPasswordReset, admin: CurrentUser = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Reset hasła użytkownika przez admina - bez wymagania starego hasła."""
    err = validate_password_strength(payload.new_password)
    if err:
        raise HTTPException(400, err)
    
    new_hash = hash_password(payload.new_password)
    r = await db.execute(text(f"UPDATE {settings.TABLE_USERS} SET password_hash = :h WHERE id = :id RETURNING email"), {"h": new_hash, "id": uid})
    u = r.first()
    await db.commit()
    if not u:
        raise HTTPException(404, "Użytkownik nie znaleziony")
    
    await log_audit(db, admin, "PASSWORD_RESET_BY_ADMIN", "user", str(uid), f"reset hasła dla: {u.email}")


@app.delete("/api/users/{uid}", status_code=204)
async def delete_user(uid: int, admin: CurrentUser = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Usuwa użytkownika - tylko admin (nie może usunąć siebie)."""
    if uid == admin.id:
        raise HTTPException(400, "Nie możesz usunąć własnego konta!")
    
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_USERS} WHERE id = :id RETURNING email"), {"id": uid})
    u = r.first()
    await db.commit()
    if not u:
        raise HTTPException(404, "Użytkownik nie znaleziony")
    
    await log_audit(db, admin, "USER_DELETED", "user", str(uid), f"usunięto: {u.email}")


# ===== AUDIT LOG (tylko admin) =====
@app.get("/api/audit-log", response_model=List[AuditLogOut])
async def get_audit_log(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    admin: CurrentUser = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Audit log - tylko super-admin (konkretny email z SUPER_ADMIN_EMAIL)."""
    # Sprawdź czy to super-admin
    super_email = settings.SUPER_ADMIN_EMAIL.strip().lower()
    if super_email and admin.email.lower() != super_email:
        raise HTTPException(403, "Audit log dostępny tylko dla super-administratora")
    
    where = []
    params = {"limit": limit, "offset": offset}
    if user_id:
        where.append("user_id = :uid")
        params["uid"] = user_id
    if action:
        where.append("action = :action")
        params["action"] = action
    
    where_clause = "WHERE " + " AND ".join(where) if where else ""
    
    r = await db.execute(text(f"""
        SELECT id, user_id, user_email, action, resource_type, resource_id, details, created_at
        FROM {settings.TABLE_AUDIT_LOG}
        {where_clause}
        ORDER BY created_at DESC
        LIMIT :limit OFFSET :offset
    """), params)
    
    return [AuditLogOut(**dict(row._mapping)) for row in r]


# ========== ENDPOINTY: META ==========
@app.get("/api/health")
async def health():
    return {"status": "ok", "time": datetime.now().isoformat(), "version": "5.0"}


@app.get("/api/stats")
async def stats(db: AsyncSession = Depends(get_db)):
    r1 = await db.execute(text(f"SELECT COUNT(*) FROM {settings.TABLE_PRODUCTS}"))
    r2 = await db.execute(text(f"SELECT COUNT(*) FROM {settings.TABLE_PRODUCTS} WHERE {settings.COL_PRODUCT_STOCK} > 0"))
    r3 = await db.execute(text(f"SELECT COUNT(*) FROM {settings.TABLE_ORDERS} WHERE {settings.COL_ORDER_DATE} >= NOW() - INTERVAL '365 days'"))
    return {"total_products": r1.scalar(), "products_with_stock": r2.scalar(), "orders_last_12m": r3.scalar()}


@app.get("/api/classification")
async def classification(db: AsyncSession = Depends(get_db)):
    products_result = await db.execute(text(SALES_QUERY), {"default_lead_time": settings.DEFAULT_LEAD_TIME_DAYS})
    counts = {"ACTIVE": 0, "ACTIVE_NO_STOCK": 0, "DEAD_STOCK": 0, "INACTIVE": 0}
    dead_stock_value = 0.0
    for r in products_result:
        row = dict(r._mapping)
        s = classify_product(row)
        counts[s] += 1
        if s == "DEAD_STOCK":
            dead_stock_value += row["stock"] * row.get("price", 0)
    return {"counts": counts, "dead_stock_value_pln": round(dead_stock_value, 2), "total": sum(counts.values())}


# ========== PRODUKTY ==========
async def _fetch_products_internal(db: AsyncSession, include_set: set) -> List[ProductSummary]:
    await auto_deliver_containers(db)
    products_result = await db.execute(text(SALES_QUERY), {"default_lead_time": settings.DEFAULT_LEAD_TIME_DAYS})
    products = [dict(r._mapping) for r in products_result]
    
    incoming_result = await db.execute(text(INCOMING_QUERY))
    incoming_all = [dict(r._mapping) for r in incoming_result]
    
    incoming_by_sku = {}
    for inc in incoming_all:
        key = inc["sku"].strip().lower() if inc["sku"] else ""
        incoming_by_sku.setdefault(key, []).append(inc)
    
    results = []
    for p in products:
        if classify_product(p) not in include_set:
            continue
        sku_key = p["sku"].strip().lower() if p["sku"] else ""
        results.append(calculate_forecast(p, incoming_by_sku.get(sku_key, [])))
    return results


@app.get("/api/products", response_model=List[ProductSummary])
async def list_products(include: str = Query("ACTIVE,ACTIVE_NO_STOCK"), db: AsyncSession = Depends(get_db)):
    allowed = set(s.strip().upper() for s in include.split(",") if s.strip())
    return await _fetch_products_internal(db, allowed)


@app.get("/api/products/{sku}", response_model=ProductSummary)
async def get_product(sku: str, db: AsyncSession = Depends(get_db)):
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"})
    for p in products:
        if p.sku == sku:
            return p
    raise HTTPException(404, f"Produkt {sku} nie znaleziony")


@app.put("/api/products/{sku}/lead-time", response_model=ProductSummary)
async def update_lead_time(sku: str, payload: LeadTimeUpdate, db: AsyncSession = Depends(get_db)):
    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_LEAD_TIMES} (sku, lead_time_days, updated_at)
            VALUES (:sku, :lt, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET lead_time_days = EXCLUDED.lead_time_days, updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "lt": payload.lead_time_days}
    )
    await db.commit()
    return await get_product(sku, db)


@app.put("/api/products/{sku}/attrs", response_model=ProductSummary)
async def update_attrs(sku: str, payload: ProductAttrsUpdate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(text(f"SELECT cbm_per_unit, manufacturer_id, seasonality_enabled, ean, forced_status FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": sku})
    e = existing.first()
    cbm = payload.cbm_per_unit if payload.cbm_per_unit is not None else (float(e.cbm_per_unit) if e else 0)
    mfr = payload.manufacturer_id if payload.manufacturer_id is not None else (e.manufacturer_id if e else None)
    seas = payload.seasonality_enabled if payload.seasonality_enabled is not None else (e.seasonality_enabled if e else False)
    ean = payload.ean if payload.ean is not None else (e.ean if e else None)
    if ean is not None and not ean.strip():
        ean = None
    
    # Forced status: "AUTO" lub "" lub null = wyczyść
    forced = payload.forced_status if payload.forced_status is not None else (e.forced_status if e else None)
    if forced in ("", "AUTO", "auto", None):
        forced = None
    elif forced not in ("ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"):
        raise HTTPException(400, f"Niepoprawny status: {forced}. Dozwolone: ACTIVE, ACTIVE_NO_STOCK, DEAD_STOCK, INACTIVE, AUTO")
    
    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, cbm_per_unit, manufacturer_id, seasonality_enabled, ean, forced_status, updated_at)
            VALUES (:sku, :cbm, :mfr, :seas, :ean, :forced, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET
                cbm_per_unit = EXCLUDED.cbm_per_unit,
                manufacturer_id = EXCLUDED.manufacturer_id,
                seasonality_enabled = EXCLUDED.seasonality_enabled,
                ean = EXCLUDED.ean,
                forced_status = EXCLUDED.forced_status,
                updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "cbm": cbm, "mfr": mfr, "seas": seas, "ean": ean, "forced": forced}
    )
    await db.commit()
    return await get_product(sku, db)


@app.get("/api/products/{sku}/projection", response_model=List[StockProjectionPoint])
async def projection(sku: str, days: int = 180, db: AsyncSession = Depends(get_db)):
    product = await get_product(sku, db)
    today = date.today()
    base_daily = product.avg_monthly_weighted / 30
    eta_map = {d.eta_date: d.quantity for d in product.incoming_deliveries}
    eta_names = {d.eta_date: f"#{d.container_number} +{d.quantity}" for d in product.incoming_deliveries}
    points = []
    current = float(product.stock)
    for offset in range(0, days + 1):
        cd = today + timedelta(days=offset)
        ev = None
        if cd in eta_map:
            current += eta_map[cd]
            ev = eta_names[cd]
        if offset > 0:
            current -= base_daily
        points.append(StockProjectionPoint(date=cd, stock=max(0, int(current)), event=ev))
    return points


@app.post("/api/products/import", response_model=ImportResult)
async def import_products(rows: List[ImportRow], db: AsyncSession = Depends(get_db)):
    """Import atrybutów dla produktów - z UI lub bezpośrednio JSON-em."""
    # Mapy istniejących SKU i producentów (case-insensitive)
    products_result = await db.execute(text(f"SELECT {settings.COL_PRODUCT_SKU} as sku FROM {settings.TABLE_PRODUCTS}"))
    valid_skus = {r._mapping["sku"].strip().lower(): r._mapping["sku"] for r in products_result}
    
    mfr_result = await db.execute(text(f"SELECT id, name FROM {settings.TABLE_MANUFACTURERS}"))
    mfr_map = {r._mapping["name"].strip().lower(): r._mapping["id"] for r in mfr_result}
    
    updated = 0
    skipped = 0
    errors = []
    
    for row in rows:
        sku_key = row.sku.strip().lower()
        if sku_key not in valid_skus:
            skipped += 1
            errors.append(f"{row.sku}: nie znaleziono w bazie")
            continue
        real_sku = valid_skus[sku_key]
        
        mfr_id = None
        if row.manufacturer_name:
            mfr_id = mfr_map.get(row.manufacturer_name.strip().lower())
            if mfr_id is None and row.manufacturer_name.strip():
                # Auto-create producent
                r = await db.execute(
                    text(f"INSERT INTO {settings.TABLE_MANUFACTURERS} (name, color) VALUES (:n, :c) ON CONFLICT (name) DO UPDATE SET name=EXCLUDED.name RETURNING id"),
                    {"n": row.manufacturer_name.strip(), "c": "#6b7280"}
                )
                mfr_id = r.scalar_one()
                mfr_map[row.manufacturer_name.strip().lower()] = mfr_id
        
        try:
            # Atrybuty
            existing = await db.execute(text(f"SELECT cbm_per_unit, manufacturer_id, seasonality_enabled FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": real_sku})
            e = existing.first()
            cbm = row.cbm if row.cbm is not None else (float(e.cbm_per_unit) if e else 0)
            new_mfr = mfr_id if mfr_id is not None else (e.manufacturer_id if e else None)
            new_seas = row.seasonality_enabled if row.seasonality_enabled is not None else (e.seasonality_enabled if e else False)
            
            await db.execute(text(f"""
                INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, cbm_per_unit, manufacturer_id, seasonality_enabled, updated_at)
                VALUES (:sku, :cbm, :mfr, :seas, CURRENT_TIMESTAMP)
                ON CONFLICT (sku) DO UPDATE SET cbm_per_unit=EXCLUDED.cbm_per_unit, manufacturer_id=EXCLUDED.manufacturer_id, seasonality_enabled=EXCLUDED.seasonality_enabled, updated_at=CURRENT_TIMESTAMP
            """), {"sku": real_sku, "cbm": cbm, "mfr": new_mfr, "seas": new_seas})
            
            # Lead time
            if row.lead_time_days is not None and 1 <= row.lead_time_days <= 365:
                await db.execute(text(f"""
                    INSERT INTO {settings.TABLE_LEAD_TIMES} (sku, lead_time_days, updated_at)
                    VALUES (:sku, :lt, CURRENT_TIMESTAMP)
                    ON CONFLICT (sku) DO UPDATE SET lead_time_days=EXCLUDED.lead_time_days, updated_at=CURRENT_TIMESTAMP
                """), {"sku": real_sku, "lt": row.lead_time_days})
            
            updated += 1
        except Exception as ex:
            errors.append(f"{row.sku}: {str(ex)[:100]}")
            skipped += 1
    
    await db.commit()
    return ImportResult(total=len(rows), updated=updated, skipped=skipped, errors=errors[:20])


@app.get("/api/products/export/csv")
async def export_xlsx(include: str = Query("ACTIVE,ACTIVE_NO_STOCK"), favorites_only: bool = Query(False), db: AsyncSession = Depends(get_db)):
    """Eksport produktów do Excela (XLSX) - polskie znaki zawsze działają."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    allowed = set(s.strip().upper() for s in include.split(",") if s.strip())
    products = await _fetch_products_internal(db, allowed)
    
    # Filtr ulubionych
    if favorites_only:
        products = [p for p in products if p.is_favorite]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Produkty"
    
    headers = [
        "SKU", "Nazwa", "EAN", "Producent", "Stan", "Cena zakupu", "Wartość PLN",
        "W drodze", "CBM", "Lead time (dni)",
        "Sprzedaż 1m", "Sprzedaż 2m", "Sprzedaż 3m", "Sprzedaż 4m",
        "YoY 30d", "YoY +30d", "Średnia miesięczna", "Miesiące zapasu",
        "Status prognozy", "Status produktu", "Sezonowy", "Obserwowany",
        "Data zamówienia", "Data końca zapasu",
    ]
    ws.append(headers)
    
    # Stylowanie nagłówka
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1c1917", end_color="1c1917", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for p in products:
        ws.append([
            p.sku, p.name, p.ean or "", p.manufacturer_name or "", p.stock,
            p.purchase_price, p.stock_value, p.stock_in_transit, p.cbm_per_unit, p.lead_time_days,
            p.sales_1m, p.sales_2m, p.sales_3m, p.sales_4m,
            p.sales_yoy_30d, p.sales_yoy_next_30d,
            p.avg_monthly_weighted, p.months_of_stock,
            p.status, p.product_status,
            "tak" if p.seasonality_enabled else "nie",
            "tak" if p.is_favorite else "nie",
            p.order_date.isoformat(), p.empty_date.isoformat(),
        ])
    
    # Auto-szerokość kolumn
    column_widths = [12, 35, 16, 15, 8, 12, 14, 10, 8, 8, 10, 10, 10, 10, 10, 10, 12, 10, 14, 14, 8, 10, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width
    
    # Zamrożenie pierwszego wiersza
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"produkty_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/containers/export/csv")
async def export_containers_xlsx(db: AsyncSession = Depends(get_db)):
    """Eksport kontenerów do Excela (XLSX)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    containers = await list_containers(db=db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontenery"
    
    headers = [
        "Nr kontenera", "Nr zamówienia", "Producent", "Typ", "Status",
        "Data zamówienia", "ETA", "SKU", "Nazwa produktu",
        "Ilość", "Cena jednostkowa", "Wartość", "CBM total",
    ]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1c1917", end_color="1c1917", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    status_label = {"ORDERED": "Zamówione", "IN_PRODUCTION": "W produkcji", "IN_TRANSIT": "W drodze", "DELIVERED": "Dostarczone"}
    
    for c in containers:
        for it in c.items:
            cena = float(it.unit_cost) if it.unit_cost else 0
            wartosc = cena * it.quantity
            ws.append([
                c.container_number, c.order_number or "",
                c.manufacturer_name or "", c.container_type_name or "",
                status_label.get(c.status, c.status),
                c.order_date.isoformat(), c.eta_date.isoformat(),
                it.sku, it.product_name or "",
                it.quantity, cena, wartosc, it.total_cbm,
            ])
    
    column_widths = [16, 16, 18, 8, 14, 14, 14, 12, 35, 8, 14, 14, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"kontenery_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ========== ANOMALIE ==========
@app.get("/api/anomalies", response_model=List[Anomaly])
async def detect_anomalies(db: AsyncSession = Depends(get_db)):
    """
    Wykrywanie anomalii - rozsądna czułość:
    - sales_spike: 1m > 1.5x średniej z poprzednich 3m, ALE 1m >= 5 (żeby nie spam dla małych)
    - sales_drop: 1m < 0.4x średniej z poprzednich 3m, ALE poprzedni miesiąc był >= 5
    - stock_drain: stan = 0 i sprzedaż 1m >= 10
    """
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK"})
    anomalies = []
    
    for p in products:
        # Średnia z miesięcy 2-4 (poprzednie 3 miesiące, pomijając aktualny)
        prev_3m_total = p.sales_2m * 2 + p.sales_3m * 3 + p.sales_4m * 4 - p.sales_1m
        # to nie idealne ale daje przybliżenie z dostępnych pól; właściwie sales_2m to średnia z 2m wstecz
        # uproszczenie: porównajmy sales_1m z sales_3m (avg z ostatnich 3m bez 1m)
        prev_avg = max(p.sales_3m, 1)  # to średnia z ostatnich 3 mies, używamy jako baseline
        
        # SPIKE
        if p.sales_1m >= 5 and p.sales_1m > prev_avg * 1.5:
            change_pct = ((p.sales_1m / prev_avg) - 1) * 100
            sev = "high" if change_pct > 100 else "medium"
            anomalies.append(Anomaly(
                sku=p.sku, name=p.name, severity=sev, type="sales_spike",
                message=f"Sprzedaż wzrosła z {prev_avg}/mies do {p.sales_1m}/mies (+{change_pct:.0f}%)",
                sales_1m=p.sales_1m, sales_3m_avg=prev_avg, change_pct=round(change_pct, 1),
            ))
        # DROP
        elif prev_avg >= 5 and p.sales_1m < prev_avg * 0.4:
            change_pct = ((p.sales_1m / prev_avg) - 1) * 100
            sev = "high" if change_pct < -70 else "medium"
            anomalies.append(Anomaly(
                sku=p.sku, name=p.name, severity=sev, type="sales_drop",
                message=f"Sprzedaż spadła z {prev_avg}/mies do {p.sales_1m}/mies ({change_pct:.0f}%)",
                sales_1m=p.sales_1m, sales_3m_avg=prev_avg, change_pct=round(change_pct, 1),
            ))
        # STOCK DRAIN
        elif p.stock == 0 and p.sales_1m >= 10 and p.stock_in_transit == 0:
            anomalies.append(Anomaly(
                sku=p.sku, name=p.name, severity="high", type="stock_drain",
                message=f"Zero stanu, sprzedaż {p.sales_1m}/mies, brak kontenera w drodze!",
                sales_1m=p.sales_1m, sales_3m_avg=prev_avg, change_pct=0,
            ))
    
    # Sortujemy: high > medium > low, a w obrębie po sprzedaży
    sev_order = {"high": 0, "medium": 1, "low": 2}
    anomalies.sort(key=lambda a: (sev_order[a.severity], -a.sales_1m))
    return anomalies[:20]


# ========== LISTA ZAKUPÓW ==========
@app.get("/api/shopping-list", response_model=List[ShoppingListGroup])
async def shopping_list(db: AsyncSession = Depends(get_db)):
    """Grupy produktów do zamówienia per producent."""
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK"})
    needing = [p for p in products if p.status in ("KRYTYCZNY", "ZAMOW_TERAZ", "ZAMOW_WKROTCE") and p.avg_monthly_weighted >= 1]
    
    # Pobierz emaile producentów
    mfr_result = await db.execute(text(f"SELECT id, name, color, email FROM {settings.TABLE_MANUFACTURERS}"))
    mfr_emails = {r._mapping["id"]: r._mapping["email"] for r in mfr_result}
    
    groups = {}
    for p in needing:
        key = p.manufacturer_id or 0
        if key not in groups:
            groups[key] = {
                "manufacturer_id": p.manufacturer_id,
                "manufacturer_name": p.manufacturer_name,
                "manufacturer_color": p.manufacturer_color,
                "manufacturer_email": mfr_emails.get(p.manufacturer_id) if p.manufacturer_id else None,
                "products": [],
                "total_skus": 0,
            }
        # Sugerowana ilość: 6 miesięcy zapasu
        recommended = max(1, int(p.avg_monthly_weighted * 6 - p.stock - p.stock_in_transit))
        groups[key]["products"].append({
            "sku": p.sku, "name": p.name,
            "stock": p.stock, "stock_in_transit": p.stock_in_transit,
            "avg_monthly": p.avg_monthly_weighted,
            "recommended_quantity": recommended,
            "purchase_price": p.purchase_price,
            "cbm_per_unit": p.cbm_per_unit,
            "status": p.status,
            "days_until_empty": p.days_until_empty,
        })
        groups[key]["total_skus"] += 1
    
    return list(groups.values())


# ========== AUTO-SUGESTIA ==========
class AutoSuggestRequest(BaseModel):
    manufacturer_id: int
    container_type_id: int
    months_horizon: int = Field(6, ge=1, le=24)


class AutoSuggestItem(BaseModel):
    sku: str
    name: str
    quantity: int
    unit_cost: float
    cbm_total: float
    is_partial: bool = False


class AutoSuggestResponse(BaseModel):
    items: List[AutoSuggestItem]
    total_cbm: float
    capacity_cbm: float
    fill_pct: float
    total_value: float
    total_units: int


@app.post("/api/auto-suggest", response_model=AutoSuggestResponse)
async def auto_suggest(payload: AutoSuggestRequest, db: AsyncSession = Depends(get_db)):
    """Algorytm proponuje skład kontenera dla danego producenta."""
    type_result = await db.execute(text(f"SELECT capacity_cbm FROM {settings.TABLE_CONTAINER_TYPES} WHERE id = :id"), {"id": payload.container_type_id})
    capacity_row = type_result.first()
    if not capacity_row:
        raise HTTPException(404, "Typ kontenera nie znaleziony")
    capacity = float(capacity_row.capacity_cbm)
    
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK"})
    mfr_products = [p for p in products if p.manufacturer_id == payload.manufacturer_id]
    mfr_products.sort(key=lambda p: (p.days_until_order, -p.avg_monthly_weighted))
    
    items = []
    used_cbm = 0.0
    
    for p in mfr_products:
        needed = int(p.avg_monthly_weighted * payload.months_horizon - p.stock - p.stock_in_transit)
        if needed <= 0:
            continue
        cbm = p.cbm_per_unit * needed
        if used_cbm + cbm > capacity:
            # Częściowo - ile się zmieści?
            remaining = capacity - used_cbm
            if p.cbm_per_unit > 0:
                fit_qty = int(remaining / p.cbm_per_unit)
                if fit_qty > 0:
                    items.append(AutoSuggestItem(
                        sku=p.sku, name=p.name, quantity=fit_qty,
                        unit_cost=p.purchase_price,
                        cbm_total=round(fit_qty * p.cbm_per_unit, 3),
                        is_partial=True,
                    ))
                    used_cbm += fit_qty * p.cbm_per_unit
            break
        items.append(AutoSuggestItem(
            sku=p.sku, name=p.name, quantity=needed,
            unit_cost=p.purchase_price,
            cbm_total=round(cbm, 3),
        ))
        used_cbm += cbm
    
    fill_pct = (used_cbm / capacity * 100) if capacity > 0 else 0
    total_value = sum(i.unit_cost * i.quantity for i in items)
    total_units = sum(i.quantity for i in items)
    
    return AutoSuggestResponse(
        items=items, total_cbm=round(used_cbm, 3),
        capacity_cbm=capacity, fill_pct=round(fill_pct, 1),
        total_value=round(total_value, 2), total_units=total_units,
    )


# ========== WYSZUKIWARKA EAN ==========
@app.get("/api/search/ean")
async def search_ean(q: str = Query(..., min_length=2), db: AsyncSession = Depends(get_db)):
    """Wyszukiwanie produktu po EAN lub SKU. Sprawdza zarówno zapisane EANy w app_product_attrs jak i historyczne w sellasist_order_items."""
    r = await db.execute(text(f"""
        SELECT DISTINCT 
            p.{settings.COL_PRODUCT_SKU} AS sku,
            p.{settings.COL_PRODUCT_NAME} AS name,
            p.{settings.COL_PRODUCT_STOCK} AS stock,
            COALESCE(pa.ean, (SELECT MAX(oi.{settings.COL_ITEM_EAN}) 
             FROM {settings.TABLE_ORDER_ITEMS} oi 
             WHERE LOWER(TRIM(oi.{settings.COL_ITEM_SKU})) = LOWER(TRIM(p.{settings.COL_PRODUCT_SKU}))
             AND oi.{settings.COL_ITEM_EAN} IS NOT NULL
             LIMIT 1)) AS ean
        FROM {settings.TABLE_PRODUCTS} p
        LEFT JOIN {settings.TABLE_PRODUCT_ATTRS} pa ON pa.sku = p.{settings.COL_PRODUCT_SKU}
        WHERE LOWER(p.{settings.COL_PRODUCT_SKU}) LIKE LOWER(:q)
           OR pa.ean LIKE :q
           OR EXISTS (
                SELECT 1 FROM {settings.TABLE_ORDER_ITEMS} oi
                WHERE LOWER(TRIM(oi.{settings.COL_ITEM_SKU})) = LOWER(TRIM(p.{settings.COL_PRODUCT_SKU}))
                  AND oi.{settings.COL_ITEM_EAN} LIKE :q
           )
        LIMIT 10
    """), {"q": f"%{q}%"})
    return [{"sku": r._mapping["sku"], "name": r._mapping["name"], "stock": r._mapping["stock"], "ean": r._mapping["ean"]} for r in r]


# ========== WYKRES WARTOŚCI MAGAZYNU ==========
@app.get("/api/stock-value-history")
async def stock_value_history(days: int = 90, db: AsyncSession = Depends(get_db)):
    """
    Symulacja wartości magazynu w czasie - bazuje na obecnym stanie + sprzedaży.
    To jest aproksymacja, bo Subiekt nie trzyma historii stanu - rekonstruujemy z danych zamówień.
    """
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK"})
    
    # Dla każdego dnia obliczamy: aktualny_stan + sprzedaż_ktora_byla_pomiedzy_dzis_a_dniem
    # Czyli: stan_w_dniu_X = stan_dzis + sumaryczna_sprzedaż_od_X_do_dzis
    today = date.today()
    
    # Pobierz sprzedaż dziennie z 90 dni wstecz
    sales_query = f"""
        SELECT 
            LOWER(TRIM(oi.{settings.COL_ITEM_SKU})) AS sku_norm,
            DATE(o.{settings.COL_ORDER_DATE}) AS sale_date,
            SUM(oi.{settings.COL_ITEM_QTY}) AS qty
        FROM {settings.TABLE_ORDER_ITEMS} oi
        JOIN {settings.TABLE_ORDERS} o ON o.{settings.COL_ORDER_ID} = oi.{settings.COL_ITEM_ORDER_ID}
        WHERE o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '{days} days'
        GROUP BY LOWER(TRIM(oi.{settings.COL_ITEM_SKU})), DATE(o.{settings.COL_ORDER_DATE})
    """
    sales_result = await db.execute(text(sales_query))
    
    # Mapa: {sku_norm: {date: qty}}
    sales_by_sku = {}
    for r in sales_result:
        sku_norm = r._mapping["sku_norm"]
        sale_date = r._mapping["sale_date"]
        qty = r._mapping["qty"]
        if sku_norm not in sales_by_sku:
            sales_by_sku[sku_norm] = {}
        sales_by_sku[sku_norm][sale_date] = qty
    
    # Mapa SKU → cena
    price_map = {p.sku.strip().lower(): p.purchase_price for p in products}
    stock_map = {p.sku.strip().lower(): p.stock for p in products}
    
    # Obliczamy wartość dla każdego dnia
    points = []
    for offset in range(days, -1, -1):
        d = today - timedelta(days=offset)
        total_value = 0
        for sku_norm in stock_map:
            stock_today = stock_map[sku_norm]
            price = price_map.get(sku_norm, 0)
            # Stan w dniu d = obecny + sprzedaż między d a dziś
            sold_between = 0
            if sku_norm in sales_by_sku:
                for sale_d, qty in sales_by_sku[sku_norm].items():
                    if sale_d > d:
                        sold_between += qty
            stock_at_d = stock_today + sold_between
            total_value += stock_at_d * price
        points.append({"date": d.isoformat(), "value": round(total_value, 2)})
    
    return {"points": points, "current_value": points[-1]["value"] if points else 0}


# ========== PRODUCENCI (CRUD) ==========
@app.get("/api/manufacturers", response_model=List[ManufacturerOut])
async def list_manufacturers(db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"SELECT id, name, color, notes, email FROM {settings.TABLE_MANUFACTURERS} ORDER BY name"))
    return [ManufacturerOut(**dict(row._mapping)) for row in r]


@app.post("/api/manufacturers", response_model=ManufacturerOut, status_code=201)
async def create_manufacturer(payload: ManufacturerIn, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        text(f"INSERT INTO {settings.TABLE_MANUFACTURERS} (name, color, notes, email) VALUES (:n, :c, :no, :e) RETURNING id"),
        {"n": payload.name, "c": payload.color, "no": payload.notes, "e": payload.email}
    )
    new_id = r.scalar_one()
    await db.commit()
    return ManufacturerOut(id=new_id, **payload.model_dump())


@app.patch("/api/manufacturers/{mid}", response_model=ManufacturerOut)
async def update_manufacturer(mid: int, payload: ManufacturerIn, db: AsyncSession = Depends(get_db)):
    await db.execute(
        text(f"UPDATE {settings.TABLE_MANUFACTURERS} SET name=:n, color=:c, notes=:no, email=:e WHERE id=:id"),
        {"n": payload.name, "c": payload.color, "no": payload.notes, "e": payload.email, "id": mid}
    )
    await db.commit()
    return ManufacturerOut(id=mid, **payload.model_dump())


@app.delete("/api/manufacturers/{mid}", status_code=204)
async def delete_manufacturer(mid: int, db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_MANUFACTURERS} WHERE id=:id"), {"id": mid})
    await db.commit()
    if r.rowcount == 0: raise HTTPException(404)


# ========== TYPY KONTENERÓW (CRUD) ==========
@app.get("/api/container-types", response_model=List[ContainerTypeOut])
async def list_container_types(db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"SELECT id, name, capacity_cbm, sort_order FROM {settings.TABLE_CONTAINER_TYPES} ORDER BY sort_order, name"))
    return [ContainerTypeOut(id=row._mapping["id"], name=row._mapping["name"],
                              capacity_cbm=float(row._mapping["capacity_cbm"]),
                              sort_order=row._mapping["sort_order"]) for row in r]


@app.post("/api/container-types", response_model=ContainerTypeOut, status_code=201)
async def create_container_type(payload: ContainerTypeIn, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        text(f"INSERT INTO {settings.TABLE_CONTAINER_TYPES} (name, capacity_cbm, sort_order) VALUES (:n, :c, :s) RETURNING id"),
        {"n": payload.name, "c": payload.capacity_cbm, "s": payload.sort_order}
    )
    new_id = r.scalar_one()
    await db.commit()
    return ContainerTypeOut(id=new_id, **payload.model_dump())


@app.patch("/api/container-types/{tid}", response_model=ContainerTypeOut)
async def update_container_type(tid: int, payload: ContainerTypeIn, db: AsyncSession = Depends(get_db)):
    await db.execute(
        text(f"UPDATE {settings.TABLE_CONTAINER_TYPES} SET name=:n, capacity_cbm=:c, sort_order=:s WHERE id=:id"),
        {"n": payload.name, "c": payload.capacity_cbm, "s": payload.sort_order, "id": tid}
    )
    await db.commit()
    return ContainerTypeOut(id=tid, **payload.model_dump())


@app.delete("/api/container-types/{tid}", status_code=204)
async def delete_container_type(tid: int, db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINER_TYPES} WHERE id=:id"), {"id": tid})
    await db.commit()
    if r.rowcount == 0: raise HTTPException(404)


# ========== KONTENERY ==========
async def _fetch_attachments(db: AsyncSession, container_id: int) -> List[AttachmentOut]:
    r = await db.execute(text(f"SELECT id, filename, file_type, file_size, uploaded_at FROM {settings.TABLE_ATTACHMENTS} WHERE container_id = :c ORDER BY uploaded_at DESC"), {"c": container_id})
    return [AttachmentOut(**dict(row._mapping)) for row in r]


@app.get("/api/containers", response_model=List[ContainerOut])
async def list_containers(status: Optional[ContainerStatus] = None, db: AsyncSession = Depends(get_db)):
    where = "WHERE c.status = :status" if status else ""
    r = await db.execute(text(f"""
        SELECT 
            c.id, c.container_number, c.order_number, c.container_type_id, c.manufacturer_id,
            c.order_date, c.eta_date, c.status, c.notes,
            ct.name AS container_type_name, ct.capacity_cbm AS container_capacity_cbm,
            m.name AS manufacturer_name, m.color AS manufacturer_color,
            ci.id AS item_id, ci.sku, ci.quantity, ci.unit_cost,
            p.{settings.COL_PRODUCT_NAME} AS product_name,
            COALESCE(pa.cbm_per_unit, 0) AS cbm_per_unit
        FROM {settings.TABLE_CONTAINERS} c
        LEFT JOIN {settings.TABLE_CONTAINER_TYPES} ct ON ct.id = c.container_type_id
        LEFT JOIN {settings.TABLE_MANUFACTURERS} m ON m.id = c.manufacturer_id
        LEFT JOIN {settings.TABLE_CONTAINER_ITEMS} ci ON ci.container_id = c.id
        LEFT JOIN {settings.TABLE_PRODUCTS} p ON p.{settings.COL_PRODUCT_SKU} = ci.sku
        LEFT JOIN {settings.TABLE_PRODUCT_ATTRS} pa ON pa.sku = ci.sku
        {where}
        ORDER BY c.eta_date DESC, c.id DESC, ci.id ASC
    """), {"status": status} if status else {})
    
    rows = [dict(row._mapping) for row in r]
    containers_dict = {}
    for row in rows:
        cid = row["id"]
        if cid not in containers_dict:
            cap = float(row["container_capacity_cbm"]) if row["container_capacity_cbm"] else None
            containers_dict[cid] = {
                "id": cid, "container_number": row["container_number"],
                "order_number": row["order_number"],
                "container_type_id": row["container_type_id"],
                "container_type_name": row["container_type_name"],
                "container_capacity_cbm": cap,
                "manufacturer_id": row["manufacturer_id"],
                "manufacturer_name": row["manufacturer_name"],
                "manufacturer_color": row["manufacturer_color"],
                "order_date": row["order_date"], "eta_date": row["eta_date"],
                "status": row["status"], "notes": row["notes"],
                "items": [], "attachments": [],
                "total_units": 0, "total_cbm": 0.0, "fill_percentage": None, "total_value": 0.0,
            }
        if row["item_id"] is not None:
            cbm_pu = float(row["cbm_per_unit"]) if row["cbm_per_unit"] else 0
            tcb = cbm_pu * row["quantity"]
            cost = float(row["unit_cost"]) if row["unit_cost"] else 0
            containers_dict[cid]["items"].append(ContainerItemOut(
                id=row["item_id"], sku=row["sku"], quantity=row["quantity"],
                unit_cost=cost if cost else None, product_name=row["product_name"],
                cbm_per_unit=cbm_pu, total_cbm=round(tcb, 3),
            ))
            containers_dict[cid]["total_units"] += row["quantity"]
            containers_dict[cid]["total_cbm"] += tcb
            containers_dict[cid]["total_value"] += cost * row["quantity"]
    
    # Załączniki dla każdego kontenera
    for cid in containers_dict:
        containers_dict[cid]["attachments"] = await _fetch_attachments(db, cid)
    
    for c in containers_dict.values():
        c["total_cbm"] = round(c["total_cbm"], 3)
        c["total_value"] = round(c["total_value"], 2)
        if c["container_capacity_cbm"] and c["container_capacity_cbm"] > 0:
            c["fill_percentage"] = round((c["total_cbm"] / c["container_capacity_cbm"]) * 100, 1)
    
    return [ContainerOut(**c) for c in containers_dict.values()]


@app.get("/api/containers/{cid}", response_model=ContainerOut)
async def get_container(cid: int, db: AsyncSession = Depends(get_db)):
    cs = await list_containers(db=db)
    for c in cs:
        if c.id == cid: return c
    raise HTTPException(404)


@app.post("/api/containers", response_model=ContainerOut, status_code=201)
async def create_container(payload: ContainerCreate, db: AsyncSession = Depends(get_db)):
    if payload.eta_date < payload.order_date:
        raise HTTPException(400, "ETA nie może być przed datą zamówienia")
    
    r = await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_CONTAINERS} 
            (container_number, order_number, container_type_id, manufacturer_id, order_date, eta_date, status, notes)
            VALUES (:n, :on, :tid, :mid, :od, :eta, :st, :no)
            RETURNING id
        """),
        {"n": payload.container_number, "on": payload.order_number,
         "tid": payload.container_type_id, "mid": payload.manufacturer_id,
         "od": payload.order_date, "eta": payload.eta_date,
         "st": payload.status, "no": payload.notes}
    )
    cid = r.scalar_one()
    
    for item in payload.items:
        await db.execute(
            text(f"INSERT INTO {settings.TABLE_CONTAINER_ITEMS} (container_id, sku, quantity, unit_cost) VALUES (:c, :s, :q, :u)"),
            {"c": cid, "s": item.sku, "q": item.quantity, "u": item.unit_cost}
        )
    
    await db.commit()
    return await get_container(cid, db)


@app.patch("/api/containers/{cid}", response_model=ContainerOut)
async def update_container(cid: int, payload: ContainerUpdate, db: AsyncSession = Depends(get_db)):
    updates = []
    params = {"id": cid}
    for field in ["container_number", "order_number", "container_type_id", "manufacturer_id", "order_date", "eta_date", "status", "notes"]:
        v = getattr(payload, field)
        if v is not None:
            updates.append(f"{field} = :{field}")
            params[field] = v
    
    if updates:
        updates.append("updated_at = CURRENT_TIMESTAMP")
        await db.execute(text(f"UPDATE {settings.TABLE_CONTAINERS} SET {', '.join(updates)} WHERE id = :id"), params)
    
    if payload.items is not None:
        await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINER_ITEMS} WHERE container_id = :cid"), {"cid": cid})
        for item in payload.items:
            await db.execute(
                text(f"INSERT INTO {settings.TABLE_CONTAINER_ITEMS} (container_id, sku, quantity, unit_cost) VALUES (:c, :s, :q, :u)"),
                {"c": cid, "s": item.sku, "q": item.quantity, "u": item.unit_cost}
            )
    
    await db.commit()
    return await get_container(cid, db)


@app.delete("/api/containers/{cid}", status_code=204)
async def delete_container(cid: int, db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINERS} WHERE id = :id"), {"id": cid})
    await db.commit()
    if r.rowcount == 0: raise HTTPException(404)


@app.post("/api/containers/{cid}/deliver", response_model=ContainerOut)
async def deliver_container(cid: int, db: AsyncSession = Depends(get_db)):
    await db.execute(text(f"UPDATE {settings.TABLE_CONTAINERS} SET status = 'DELIVERED', updated_at = CURRENT_TIMESTAMP WHERE id = :id"), {"id": cid})
    await db.commit()
    return await get_container(cid, db)


# ========== ZAŁĄCZNIKI (mock metadane) ==========
@app.post("/api/containers/{cid}/attachments", response_model=AttachmentOut, status_code=201)
async def add_attachment(cid: int, payload: AttachmentCreate, db: AsyncSession = Depends(get_db)):
    """Dodaje metadane załącznika - bez prawdziwego uploadu pliku (lokalnie nie ma sensu)."""
    r = await db.execute(text(f"""
        INSERT INTO {settings.TABLE_ATTACHMENTS} (container_id, filename, file_type, file_size)
        VALUES (:c, :n, :t, :s) RETURNING id, uploaded_at
    """), {"c": cid, "n": payload.filename, "t": payload.file_type, "s": payload.file_size})
    row = r.first()
    await db.commit()
    return AttachmentOut(id=row.id, filename=payload.filename, file_type=payload.file_type, file_size=payload.file_size, uploaded_at=row.uploaded_at)


@app.delete("/api/attachments/{aid}", status_code=204)
async def delete_attachment(aid: int, db: AsyncSession = Depends(get_db)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_ATTACHMENTS} WHERE id = :id"), {"id": aid})
    await db.commit()
    if r.rowcount == 0: raise HTTPException(404)


# ========== KALENDARZ ==========
@app.get("/api/calendar")
async def calendar_events(db: AsyncSession = Depends(get_db)):
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK"})
    containers = await list_containers(db=db)
    
    events = []
    for p in products:
        if p.status in ("KRYTYCZNY", "ZAMOW_TERAZ", "ZAMOW_WKROTCE") and p.avg_monthly_weighted >= 1:
            events.append({"date": p.order_date.isoformat(), "type": "ORDER", "sku": p.sku, "name": p.name, "status": p.status})
            events.append({"date": p.empty_date.isoformat(), "type": "EMPTY", "sku": p.sku, "name": p.name, "status": p.status})
    
    for c in containers:
        if c.status != "DELIVERED":
            events.append({
                "date": c.eta_date.isoformat(), "type": "DELIVERY",
                "container_id": c.id, "container_number": c.container_number,
                "order_number": c.order_number, "manufacturer_name": c.manufacturer_name,
                "manufacturer_color": c.manufacturer_color, "total_units": c.total_units,
                "container_status": c.status,
            })
    
    return events


# ========== CASHFLOW ==========
@app.get("/api/cashflow")
async def cashflow(months: int = 6, db: AsyncSession = Depends(get_db)):
    containers = await list_containers(db=db)
    today = date.today()
    
    result = []
    for i in range(months):
        year = today.year + (today.month - 1 + i) // 12
        month = (today.month - 1 + i) % 12 + 1
        result.append({
            "year": year, "month": month,
            "label": date(year, month, 1).strftime("%Y-%m"),
            "containers": [], "total": 0.0,
        })
    
    for c in containers:
        if c.status == "DELIVERED": continue
        eta = c.eta_date
        idx = (eta.year - today.year) * 12 + (eta.month - today.month)
        if 0 <= idx < months:
            result[idx]["containers"].append({
                "id": c.id, "container_number": c.container_number,
                "order_number": c.order_number,
                "manufacturer_name": c.manufacturer_name,
                "manufacturer_color": c.manufacturer_color,
                "eta_date": c.eta_date.isoformat(),
                "total_value": c.total_value,
            })
            result[idx]["total"] += c.total_value
    
    return {"months": result, "total": round(sum(m["total"] for m in result), 2)}


# ========== ULUBIONE ==========
@app.put("/api/products/{sku}/favorite", response_model=ProductSummary)
async def toggle_favorite(sku: str, db: AsyncSession = Depends(get_db)):
    """Przełącza status ulubione - jeśli był true, robi false i odwrotnie."""
    # Sprawdź czy istnieje wiersz
    existing = await db.execute(text(f"SELECT is_favorite FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": sku})
    e = existing.first()
    new_val = not e.is_favorite if e else True
    
    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, is_favorite, updated_at)
            VALUES (:sku, :fav, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET is_favorite = EXCLUDED.is_favorite, updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "fav": new_val}
    )
    await db.commit()
    return await get_product(sku, db)


@app.get("/api/favorites", response_model=List[ProductSummary])
async def list_favorites(db: AsyncSession = Depends(get_db)):
    """Zwraca tylko ulubione produkty."""
    products = await _fetch_products_internal(db, {"ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"})
    return [p for p in products if p.is_favorite]


# ========== GLOBALNA WYSZUKIWARKA ==========
@app.get("/api/search/global")
async def search_global(q: str = Query(..., min_length=2), db: AsyncSession = Depends(get_db)):
    """
    Globalna wyszukiwarka po: SKU, nazwie produktu, EAN, producencie, numerze kontenera.
    Zwraca wyniki pogrupowane.
    """
    query_lower = f"%{q.lower()}%"
    
    # 1. Produkty (SKU + nazwa)
    products_result = await db.execute(text(f"""
        SELECT 
            p.{settings.COL_PRODUCT_SKU} AS sku,
            p.{settings.COL_PRODUCT_NAME} AS name,
            COALESCE(p.{settings.COL_PRODUCT_STOCK}, 0) AS stock,
            m.name AS manufacturer_name,
            m.color AS manufacturer_color
        FROM {settings.TABLE_PRODUCTS} p
        LEFT JOIN {settings.TABLE_PRODUCT_ATTRS} pa ON pa.sku = p.{settings.COL_PRODUCT_SKU}
        LEFT JOIN {settings.TABLE_MANUFACTURERS} m ON m.id = pa.manufacturer_id
        WHERE LOWER(p.{settings.COL_PRODUCT_SKU}) LIKE :q 
           OR LOWER(p.{settings.COL_PRODUCT_NAME}) LIKE :q
        ORDER BY 
            CASE WHEN LOWER(p.{settings.COL_PRODUCT_SKU}) = LOWER(:exact) THEN 0 ELSE 1 END,
            p.{settings.COL_PRODUCT_SKU}
        LIMIT 15
    """), {"q": query_lower, "exact": q})
    products = [dict(r._mapping) for r in products_result]
    
    # 2. Wyszukiwanie po EAN (jeśli zapytanie wygląda jak liczba)
    ean_products = []
    if q.replace(" ", "").isdigit() or any(c.isdigit() for c in q):
        ean_result = await db.execute(text(f"""
            SELECT DISTINCT 
                oi.{settings.COL_ITEM_SKU} AS sku,
                oi.{settings.COL_ITEM_EAN} AS ean,
                MAX(p.{settings.COL_PRODUCT_NAME}) AS name
            FROM {settings.TABLE_ORDER_ITEMS} oi
            LEFT JOIN {settings.TABLE_PRODUCTS} p 
                ON LOWER(TRIM(p.{settings.COL_PRODUCT_SKU})) = LOWER(TRIM(oi.{settings.COL_ITEM_SKU}))
            WHERE oi.{settings.COL_ITEM_EAN} LIKE :q
            GROUP BY oi.{settings.COL_ITEM_SKU}, oi.{settings.COL_ITEM_EAN}
            LIMIT 10
        """), {"q": query_lower})
        ean_products = [dict(r._mapping) for r in ean_result]
    
    # 3. Producenci
    mfrs_result = await db.execute(text(f"""
        SELECT id, name, color, email, notes
        FROM {settings.TABLE_MANUFACTURERS}
        WHERE LOWER(name) LIKE :q OR LOWER(COALESCE(notes, '')) LIKE :q OR LOWER(COALESCE(email, '')) LIKE :q
        LIMIT 10
    """), {"q": query_lower})
    manufacturers = [dict(r._mapping) for r in mfrs_result]
    
    # 4. Kontenery
    containers_result = await db.execute(text(f"""
        SELECT 
            c.id, c.container_number, c.order_number, c.eta_date, c.status,
            m.name AS manufacturer_name, m.color AS manufacturer_color
        FROM {settings.TABLE_CONTAINERS} c
        LEFT JOIN {settings.TABLE_MANUFACTURERS} m ON m.id = c.manufacturer_id
        WHERE LOWER(c.container_number) LIKE :q 
           OR LOWER(COALESCE(c.order_number, '')) LIKE :q
           OR LOWER(COALESCE(c.notes, '')) LIKE :q
        ORDER BY c.eta_date DESC
        LIMIT 10
    """), {"q": query_lower})
    containers = [dict(r._mapping) for r in containers_result]
    
    return {
        "products": products,
        "ean": ean_products,
        "manufacturers": manufacturers,
        "containers": containers,
        "total": len(products) + len(ean_products) + len(manufacturers) + len(containers),
    }


# ========== PDF GENERATOR (dane do PDF, generowane po stronie frontendu) ==========
class OrderPdfRequest(BaseModel):
    manufacturer_id: int
    items: List[dict]  # [{sku, name, quantity, unit_cost}]
    notes: Optional[str] = None
    custom_order_number: Optional[str] = None


@app.post("/api/order-pdf-data")
async def order_pdf_data(payload: OrderPdfRequest, db: AsyncSession = Depends(get_db)):
    """
    Zwraca dane do wygenerowania PDF zamówienia.
    PDF generujemy po stronie frontendu (jsPDF) bo łatwiej kontrolować layout.
    """
    mfr = await db.execute(text(f"SELECT id, name, email, notes, color FROM {settings.TABLE_MANUFACTURERS} WHERE id = :id"), {"id": payload.manufacturer_id})
    m = mfr.first()
    if not m:
        raise HTTPException(404, "Producent nie znaleziony")
    
    total_value = sum(i.get("quantity", 0) * i.get("unit_cost", 0) for i in payload.items)
    total_units = sum(i.get("quantity", 0) for i in payload.items)
    
    return {
        "manufacturer": {"id": m.id, "name": m.name, "email": m.email, "notes": m.notes, "color": m.color},
        "order_number": payload.custom_order_number or f"PO-{date.today().strftime('%Y%m%d')}-{m.id:03d}",
        "order_date": date.today().isoformat(),
        "items": payload.items,
        "total_value": round(total_value, 2),
        "total_units": total_units,
        "notes": payload.notes,
    }

