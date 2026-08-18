"""Raporty: zbiorczy (KPI w czasie), per SKU i zajętość magazynu (m³).

Dane pochodzą ze snapshotów zbieranych 2× dziennie (7:05 / 20:05), więc są DOKŁADNE
— nie rekonstruujemy przeszłości. Okres bez snapshotów zwraca pustkę, nie zmyślone liczby.

Zajętość magazynu liczona jest zawsze NA ŻYWO (nie ze snapshotów) i siedzi za osobnym
uprawnieniem `viewOccupancy` — patrz sekcja na dole pliku.
"""
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from config import settings, INCLUDED_STATUS_FILTER
from sql import PRODUCT_NAMES_QUERY
from database import get_db
from models import CurrentUser
from security import get_current_user, has_perm, resolve_scope
from services.containers import fetch_containers
from services.snapshots import store_snapshot, build_kpi_rows, build_stock_rows, SLOTS
from audit import log_audit

router = APIRouter(prefix="/api", tags=["reports"])

# Klucze celowo zostają przy starych nazwach kolumn w tabeli snapshotów — zmieniamy TREŚĆ,
# którą niosą, i etykiety, a nie schemat bazy (nie chcemy migracji dla zmiany nazewnictwa):
#   zaplacono_pln      → zapłacone zaliczki+balance ZIELONYCH kontenerów  (pulpit: „Magazyn w drodze")
#   do_zaplacenia_pln  → niezapłacone kwoty tych samych ZIELONYCH         (pulpit: podpis „do zapłacenia")
#   pozostalo_pln      → niezapłacone kwoty CZERWONYCH                    (pulpit: „W Prognozie")
# Kolumny magazyn_w_drodze_pln i kontenery_pln nadal się zapisują (wartość towaru z ERP
# i wartość czerwonych lotów), tylko nie są już pokazywane — liczyły co innego niż pulpit.
KPI_FIELDS = [
    ("kapital_pln", "Kapitał w towarze"),
    ("magazyn_pln", "Wartość magazynu"),
    ("zaplacono_pln", "Magazyn w drodze — opłacone"),
    ("do_zaplacenia_pln", "Magazyn w drodze — do zapłaty"),
    ("pozostalo_pln", "W prognozie"),
]
# 'wieczor' ma pierwszeństwo — to „stan na koniec dnia".
SLOT_ORDER = "CASE snap_slot WHEN 'wieczor' THEN 0 ELSE 1 END"


def _require_reports(user: CurrentUser = Depends(get_current_user)) -> CurrentUser:
    if not has_perm(user, "viewReports"):
        raise HTTPException(403, "Brak uprawnienia do raportów")
    return user


def _parse_day(s: str, field: str) -> date:
    try:
        return date.fromisoformat(s)
    except Exception:
        raise HTTPException(400, f"{field}: data w formacie RRRR-MM-DD")


def _range(date_from: str, date_to: str) -> tuple:
    a = _parse_day(date_from, "from")
    b = _parse_day(date_to, "to") if date_to else a
    if b < a:
        a, b = b, a
    return a, b


# ── snapshot ręczny ──────────────────────────────────────────

@router.post("/reports/snapshot")
async def make_snapshot(slot: str = Query("wieczor"), db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports)):
    """Wymusza zapis snapshotu teraz (poza harmonogramem). Idempotentnie — nadpisuje tę porę."""
    if slot not in SLOTS:
        raise HTTPException(400, f"Pora musi być jedną z: {', '.join(SLOTS)}")
    return await store_snapshot(db, slot)


@router.get("/reports/available")
async def available(db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports)):
    """Zakres dat, dla których w ogóle są snapshoty (do ustawienia fragmentatora)."""
    r = await db.execute(text(f"SELECT MIN(snap_date), MAX(snap_date), COUNT(DISTINCT snap_date) FROM {settings.TABLE_KPI_SNAPSHOTS}"))
    row = r.first()
    return {
        "first": row[0].isoformat() if row and row[0] else None,
        "last": row[1].isoformat() if row and row[1] else None,
        "days": int(row[2] or 0) if row else 0,
    }


# ── raport zbiorczy (KPI) ────────────────────────────────────

@router.get("/reports/kpi-range")
async def kpi_range(
    date_from: str = Query(..., alias="from"), date_to: str = Query("", alias="to"),
    scope: str = Query("all"), group: str = Query("day"), slot: str = Query(""),
    db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports),
):
    """Serie KPI w okresie.

    group="day"   → jeden wiersz na dzień (domyślnie snapshot wieczorny = stan na koniec dnia)
    group="month" → jeden wiersz na miesiąc (ostatni snapshot miesiąca)
    slot="rano"/"wieczor" → wymusza konkretną porę zamiast „ostatniej z dnia".
    """
    # Zakres firmowy usera wygrywa nad parametrem z frontu — scoped user nie dostanie "all".
    scope = resolve_scope(scope, user)
    a, b = _range(date_from, date_to)
    params = {"f": scope, "a": a, "b": b}
    slot_where = ""
    if slot:
        if slot not in SLOTS:
            raise HTTPException(400, "Nieznana pora")
        slot_where = " AND snap_slot = :slot"
        params["slot"] = slot

    if group == "month":
        key_expr = "date_trunc('month', snap_date)"
        label_expr = "to_char(snap_date, 'YYYY-MM')"
    else:
        key_expr = "snap_date"
        label_expr = "to_char(snap_date, 'YYYY-MM-DD')"

    # Jeden dzień bez wymuszonej pory → pokazujemy OBIE pory, żeby było widać ruch w ciągu dnia.
    intraday = (a == b and not slot and group != "month")
    if intraday:
        r = await db.execute(text(f"""
            SELECT to_char(snap_date, 'YYYY-MM-DD') || ' · ' ||
                   CASE snap_slot WHEN 'rano' THEN 'rano' ELSE 'wieczór' END AS label,
                   snap_date, snap_slot,
                   kapital_pln, magazyn_pln, magazyn_w_drodze_pln, kontenery_pln,
                   zaplacono_pln, pozostalo_pln, do_zaplacenia_pln
            FROM {settings.TABLE_KPI_SNAPSHOTS}
            WHERE firma_slug = :f AND snap_date = :a
            ORDER BY CASE snap_slot WHEN 'rano' THEN 0 ELSE 1 END
        """), params)
    else:
        r = await db.execute(text(f"""
            SELECT DISTINCT ON ({key_expr})
                   {label_expr} AS label, snap_date, snap_slot,
                   kapital_pln, magazyn_pln, magazyn_w_drodze_pln, kontenery_pln,
                   zaplacono_pln, pozostalo_pln, do_zaplacenia_pln
            FROM {settings.TABLE_KPI_SNAPSHOTS}
            WHERE firma_slug = :f AND snap_date >= :a AND snap_date <= :b{slot_where}
            ORDER BY {key_expr}, snap_date DESC, {SLOT_ORDER}
        """), params)

    rows = []
    for row in r:
        d = dict(row._mapping)
        d["snap_date"] = d["snap_date"].isoformat()
        for k, _ in KPI_FIELDS:
            d[k] = float(d[k] or 0)
        rows.append(d)
    rows.sort(key=lambda x: x["label"])

    first, last = (rows[0], rows[-1]) if rows else (None, None)
    summary = []
    for key, label in KPI_FIELDS:
        s_val = first[key] if first else None
        e_val = last[key] if last else None
        delta = round(e_val - s_val, 2) if (s_val is not None and e_val is not None) else None
        delta_pct = round(((e_val - s_val) / s_val) * 100, 1) if (s_val not in (None, 0) and e_val is not None) else None
        summary.append({"key": key, "label": label, "start": s_val, "end": e_val,
                        "delta": delta, "delta_pct": delta_pct})

    return {"from": a.isoformat(), "to": b.isoformat(), "scope": scope, "group": group,
            "has_data": bool(rows), "rows": rows, "summary": summary,
            "fields": [{"key": k, "label": l} for k, l in KPI_FIELDS]}


# ── raport per SKU ───────────────────────────────────────────

async def _sku_snapshot(db: AsyncSession, day_lo: date, day_hi: date, newest: bool, slot: str):
    """Jeden snapshot per SKU: najnowszy (newest=True) albo najstarszy w oknie dat."""
    params = {"a": day_lo, "b": day_hi}
    slot_where = ""
    if slot:
        slot_where = " AND snap_slot = :slot"
        params["slot"] = slot
    order_dir = "DESC" if newest else "ASC"
    slot_dir = SLOT_ORDER if newest else f"CASE snap_slot WHEN 'rano' THEN 0 ELSE 1 END"
    r = await db.execute(text(f"""
        SELECT DISTINCT ON (sku, firma_slug)
               sku, nazwa, firma_slug, cena_jednostkowa,
               stan_glowny, stan_w_drodze, w_kontenerze, snap_date, snap_slot
        FROM {settings.TABLE_STOCK_SNAPSHOTS}
        WHERE snap_date >= :a AND snap_date <= :b{slot_where}
        ORDER BY sku, firma_slug, snap_date {order_dir}, {slot_dir}
    """), params)
    out = {}
    for row in r:
        d = dict(row._mapping)
        d["snap_date"] = d["snap_date"].isoformat()
        out[((d["sku"] or "").upper(), d["firma_slug"] or "amh")] = d
    return out



def _collapse(items: List[dict], scope: str) -> List[dict]:
    """Zakres = konkretna firma → filtr. „Wszyscy" → jeden wiersz na SKU, stany zsumowane
    ze wszystkich magazynów (jak w Produktach: 11 na AMH + 254 na Veluxie = 265)."""
    if scope != "all":
        return [r for r in items if (r.get("firma_slug") or "") == scope]
    merged: dict = {}
    for r in items:
        key = (r["sku"] or "").upper()
        m = merged.get(key)
        if m is None:
            m = merged[key] = dict(r)
            m["_firmy"] = set()
        else:
            for f in ("stan_glowny", "stan_w_drodze", "w_kontenerze", "razem",
                      "razem_start", "razem_end", "delta_szt"):
                if f in m and f in r and m[f] is not None and r[f] is not None:
                    m[f] += r[f]
            if not m.get("nazwa"):
                m["nazwa"] = r.get("nazwa")
            if not m.get("cena_jednostkowa"):
                m["cena_jednostkowa"] = r.get("cena_jednostkowa")
        if r.get("firma_slug"):
            m["_firmy"].add(r["firma_slug"])
    out = []
    for m in merged.values():
        firmy = sorted(m.pop("_firmy", set()))
        m["firma_slug"] = "+".join(f.upper() for f in firmy) if firmy else ""
        out.append(m)
    return out


@router.get("/reports/sku")
async def sku_report(
    date_from: str = Query(..., alias="from"), date_to: str = Query("", alias="to"),
    favorites_only: bool = Query(False), skus: str = Query(""), slot: str = Query(""),
    scope: str = Query("all"),
    db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports),
):
    """Stany per SKU.

    Jeden dzień  → zdjęcie z tego dnia.
    Zakres dat   → początek vs koniec + zmiana (tryb „b”), jeden wiersz na SKU.
    Filtry: tylko ulubione oraz ręczna lista SKU (przecinkami).
    """
    scope = resolve_scope(scope, user)
    a, b = _range(date_from, date_to)
    is_range = b > a

    end = await _sku_snapshot(db, a, b, newest=True, slot=slot)
    start: dict = {}
    compare = "none"
    if is_range:
        start = await _sku_snapshot(db, a, b, newest=False, slot=slot)
        compare = "range"
    elif not slot:
        # Ten sam dzień: jeśli są obie pory, porównujemy ranek z wieczorem (ruch dzienny).
        rano = await _sku_snapshot(db, a, b, newest=False, slot="rano")
        wieczor = await _sku_snapshot(db, a, b, newest=True, slot="wieczor")
        if rano and wieczor:
            start, end, compare = rano, wieczor, "intraday"
    has_compare = compare != "none"

    # ulubione — flaga trzymana przy atrybutach produktu
    favs = set()
    if favorites_only:
        rf = await db.execute(text(f"SELECT UPPER(TRIM(sku)) FROM {settings.TABLE_PRODUCT_ATTRS} WHERE COALESCE(is_favorite, FALSE) = TRUE"))
        favs = {row[0] for row in rf if row[0]}

    picked = {s.strip().upper() for s in skus.split(",") if s.strip()}

    rows = []
    for (key, firma), e in end.items():
        if favorites_only and key not in favs:
            continue
        if picked and key not in picked:
            continue
        cena = float(e["cena_jednostkowa"] or 0)
        gl, wd, kn = int(e["stan_glowny"] or 0), int(e["stan_w_drodze"] or 0), int(e["w_kontenerze"] or 0)
        razem = gl + wd + kn
        row = {
            "sku": e["sku"], "nazwa": e.get("nazwa") or "", "firma_slug": e.get("firma_slug") or "",
            "cena_jednostkowa": round(cena, 2),
            "stan_glowny": gl, "stan_w_drodze": wd, "w_kontenerze": kn,
            "razem": razem,
            "snap_date": e["snap_date"], "snap_slot": e["snap_slot"],
        }
        if has_compare:
            s = start.get((key, firma))
            s_razem = (int(s["stan_glowny"] or 0) + int(s["stan_w_drodze"] or 0) + int(s["w_kontenerze"] or 0)) if s else 0
            s_cena = float(s["cena_jednostkowa"] or 0) if s else cena
            row.update({"razem_start": s_razem, "razem_end": razem, "delta_szt": razem - s_razem})
        rows.append(row)

    rows = _collapse(rows, scope)
    rows.sort(key=lambda x: x["razem"], reverse=True)
    totals = {
        "sku_count": len(rows),
        "units": sum(r["razem"] for r in rows),
        "units_glowny": sum(r["stan_glowny"] for r in rows),
        "units_w_drodze": sum(r["stan_w_drodze"] for r in rows),
        "units_kontener": sum(r["w_kontenerze"] for r in rows),
    }
    if has_compare:
        totals["delta_szt"] = sum(r.get("delta_szt", 0) for r in rows)

    return {"from": a.isoformat(), "to": b.isoformat(), "is_range": has_compare,
            "compare": compare, "has_data": bool(rows), "rows": rows, "totals": totals}


# ── tryb LIVE (stan na teraz, bez zapisu do bazy) ────────────
# Te same funkcje, których używa pętla snapshotów — tylko bez INSERT-a.
# Dzięki temu „zdjęcie na teraz" nie zaśmieca historii.

async def _live_kpi(db: AsyncSession, scope: str) -> dict:
    rows = await build_kpi_rows(db)
    mine = next((r for r in rows if r["firma_slug"] == scope), None)
    fields = [{"key": k, "label": l} for k, l in KPI_FIELDS]
    if not mine:
        return {"from": "teraz", "to": "teraz", "scope": scope, "group": "live",
                "live": True, "has_data": False, "rows": [], "summary": [], "fields": fields}
    row = {"label": "Teraz", "snap_date": date.today().isoformat(), "snap_slot": "live"}
    for k, _ in KPI_FIELDS:
        row[k] = float(mine.get(k) or 0)
    summary = [{"key": k, "label": l, "start": None, "end": float(mine.get(k) or 0),
                "delta": None, "delta_pct": None} for k, l in KPI_FIELDS]
    return {"from": "teraz", "to": "teraz", "scope": scope, "group": "live", "live": True,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "has_data": True, "rows": [row], "summary": summary, "fields": fields}


async def _live_sku(db: AsyncSession, favorites_only: bool, skus: str, scope: str = "all") -> dict:
    src = await build_stock_rows(db)
    favs = set()
    if favorites_only:
        rf = await db.execute(text(f"SELECT UPPER(TRIM(sku)) FROM {settings.TABLE_PRODUCT_ATTRS} WHERE COALESCE(is_favorite, FALSE) = TRUE"))
        favs = {r[0] for r in rf if r[0]}
    picked = {x.strip().upper() for x in skus.split(",") if x.strip()}

    rows = []
    for e in src:
        key = (e["sku"] or "").strip().upper()
        if favorites_only and key not in favs:
            continue
        if picked and key not in picked:
            continue
        cena = float(e["cena_jednostkowa"] or 0)
        gl, wd, kn = int(e["stan_glowny"] or 0), int(e["stan_w_drodze"] or 0), int(e["w_kontenerze"] or 0)
        razem = gl + wd + kn
        rows.append({
            "sku": e["sku"], "nazwa": e.get("nazwa") or "", "firma_slug": e.get("firma_slug") or "",
            "cena_jednostkowa": round(cena, 2), "stan_glowny": gl, "stan_w_drodze": wd, "w_kontenerze": kn,
            "razem": razem, "wartosc_pln": round(razem * cena, 2),
            "snap_date": date.today().isoformat(), "snap_slot": "live",
        })
    # Zakres firmy — dokładnie jak w raporcie z historii: konkretna firma filtruje,
    # „Wszyscy" scala SKU w jeden wiersz i sumuje stany. Wcześniej `scope` był
    # przyjmowany i ignorowany, więc przełącznik firm nie działał w trybie „Teraz",
    # a przy „Wszyscy" ten sam SKU pojawiał się osobno dla każdego magazynu.
    rows = _collapse(rows, scope)
    for r in rows:
        r["wartosc_pln"] = round(r["razem"] * float(r.get("cena_jednostkowa") or 0), 2)
    rows.sort(key=lambda x: x["wartosc_pln"], reverse=True)
    return {"from": "teraz", "to": "teraz", "is_range": False, "compare": "none", "live": True,
            "has_data": bool(rows), "rows": rows,
            "totals": {"sku_count": len(rows), "units": sum(r["razem"] for r in rows),
                       "value_pln": round(sum(r["wartosc_pln"] for r in rows), 2)}}


@router.get("/reports/live/kpi")
async def live_kpi(scope: str = Query("all"), db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports)):
    """Zbiorczy stan NA TERAZ — liczony na żywo, nic nie zapisuje."""
    return await _live_kpi(db, resolve_scope(scope, user))


@router.get("/reports/live/sku")
async def live_sku(favorites_only: bool = Query(False), skus: str = Query(""), scope: str = Query("all"),
                   db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports)):
    """Stany per SKU NA TERAZ — liczone na żywo, nic nie zapisuje."""
    return await _live_sku(db, favorites_only, skus, resolve_scope(scope, user))


# ══════════════════════════════════════════════════════════════
# ZAJĘTOŚĆ MAGAZYNU (m³)
#
# Ile miejsca zajmuje towar dziś — i ile będzie zajmował w dniu X, po
# rozładowaniu kontenerów, które do tego dnia dojadą.
#
# Uprawnienie `viewOccupancy` CELOWO nie istnieje w ROLE_PERMS (security.py),
# więc ma je wyłącznie user z jawnym wpisem w kolumnie `permissions`.
#
# Rozłączność objętości (nic nie liczy się dwa razy):
#   · na stanie  = magazyn GŁÓWNY (AMH→Subiekt, Acti/Veluxa→Sellasist) — towar
#                  fizycznie stojący w hali,
#   · sprzedaż   = przy horyzoncie > 0 odejmujemy przewidywany rozchód: średnia dzienna
#                  z ostatnich 90 dni × liczba dni. Liczona z zamówień już w bazie
#                  (sellasist_orders), więc nie wołamy API Sellasista.
#   · w drodze   = pozycje niedostarczonych kontenerów z datą wejścia na magazyn
#                  PO dzisiaj. Magazyn „w drodze" z ERP (drugi magazyn Subiektu /
#                  Fakturownia) świadomie POMIJAMY: to ten sam towar co w kontenerach,
#                  tylko wbity księgowo, a on nie zajmuje jeszcze miejsca w hali.
#
# Data wejścia na magazyn = warehouse_delivery_date z services/containers.py:
#   delivered_date → expected_delivery_date → ETA + CONTAINER_CUSTOMS_DAYS (7).
# Ta sama reguła, co w Kalendarzu i Prognozie — jedna prawda o dacie.
# ══════════════════════════════════════════════════════════════

OCC_DEFAULT_CAPS = {"amh": 800.0, "acti": 600.0, "veluxa": 200.0}
OCC_FIRMA_LABELS = {"amh": "AMH", "acti": "Acti", "veluxa": "Veluxa"}
OCC_TONES = {"ok", "warning", "critical", "info", "pending", "anomaly", "accent"}
OCC_MAX_HORIZON = 365

OCC_DEFAULT_THRESHOLDS = {
    "product": [("W normie", 0.0, "ok"), ("Dużo miejsca", 5.0, "warning"), ("Za dużo", 10.0, "critical")],
    "fill":    [("Luźno", 0.0, "info"), ("Optymalnie", 55.0, "ok"), ("Ciasno", 85.0, "warning"), ("Przepełniony", 100.0, "critical")],
}


def _require_occupancy(user: CurrentUser = Depends(get_current_user)) -> CurrentUser:
    if not has_perm(user, "viewOccupancy"):
        raise HTTPException(403, "Brak uprawnienia do raportu zajętości magazynu")
    return user


class OccThresholdIn(BaseModel):
    label: str = Field(min_length=1, max_length=40)
    from_pct: float = 0.0
    tone: str = "info"


class OccConfigIn(BaseModel):
    caps: Dict[str, float] = {}
    product: List[OccThresholdIn] = []
    fill: List[OccThresholdIn] = []


def _occ_defaults(kind: str) -> List[dict]:
    return [{"label": l, "from_pct": p, "tone": t} for l, p, t in OCC_DEFAULT_THRESHOLDS[kind]]


async def _occ_caps(db: AsyncSession) -> Dict[str, float]:
    """Pojemności hal. Brak tabeli lub pusto → wartości domyślne (nie wywalamy raportu)."""
    caps = dict(OCC_DEFAULT_CAPS)
    try:
        r = await db.execute(text("SELECT LOWER(firma_slug) AS slug, capacity_m3 FROM app_warehouse_capacity"))
        found = {m["slug"]: float(m["capacity_m3"] or 0) for m in r.mappings() if m["slug"]}
        if found:
            caps.update(found)
    except Exception:
        pass
    return caps


async def _occ_thresholds(db: AsyncSession) -> Dict[str, List[dict]]:
    """Progi z bazy; pusto → domyślne. Zawsze posortowane rosnąco po from_pct."""
    out = {"product": [], "fill": []}
    try:
        r = await db.execute(text("""
            SELECT kind, label, from_pct, tone
            FROM app_occupancy_thresholds
            ORDER BY kind, from_pct
        """))
        for m in r.mappings():
            k = (m["kind"] or "").strip().lower()
            if k in out:
                out[k].append({"label": m["label"], "from_pct": float(m["from_pct"] or 0), "tone": m["tone"] or "info"})
    except Exception:
        pass
    for k in ("product", "fill"):
        if not out[k]:
            out[k] = _occ_defaults(k)
        out[k].sort(key=lambda t: t["from_pct"])
    return out


def _occ_match(thresholds: List[dict], value: float) -> dict:
    """Etykieta = najwyższy próg, który wartość przekroczyła."""
    hit = thresholds[0] if thresholds else {"label": "—", "from_pct": 0.0, "tone": "info"}
    for t in thresholds:
        if value >= t["from_pct"]:
            hit = t
    return hit


async def _occ_compute(db: AsyncSession, scope: str, horizon: int, include_sales: bool = True) -> dict:
    scope = (scope or "all").strip().lower()
    horizon = max(0, min(int(horizon or 0), OCC_MAX_HORIZON))
    today = date.today()
    cutoff = today + timedelta(days=horizon)

    caps = await _occ_caps(db)
    thresholds = await _occ_thresholds(db)
    top_product = thresholds["product"][-1] if thresholds["product"] else {"from_pct": 10.0, "label": "Za dużo", "tone": "critical"}

    # ── kubatura + przypisanie firmy z atrybutów produktu ────
    cbm_by_sku: Dict[str, float] = {}
    firma_of: Dict[str, str] = {}
    r = await db.execute(text(f"""
        SELECT UPPER(TRIM(a.sku)) AS sku,
               COALESCE(a.cbm_per_unit, 0) AS cbm,
               LOWER(COALESCE(f.slug, 'amh')) AS slug
        FROM {settings.TABLE_PRODUCT_ATTRS} a
        LEFT JOIN {settings.TABLE_FIRMY} f ON f.id = a.firma_id
        WHERE a.sku IS NOT NULL AND TRIM(a.sku) <> ''
    """))
    for m in r.mappings():
        cbm_by_sku[m["sku"]] = float(m["cbm"] or 0)
        firma_of[m["sku"]] = (m["slug"] or "amh")

    # ── nazwy: jedno źródło dla całego systemu (sql.PRODUCT_NAMES_QUERY) ──
    # Wcześniej były tu trzy osobne zapytania nadpisujące się w pętli, bez Fakturowni
    # i bez starego Subiektu — przez co część SKU Acti/Veluxa nie miała nazwy w raporcie,
    # mimo że w module Produkty nazwa była widoczna.
    names: Dict[str, str] = {}
    try:
        r = await db.execute(text(PRODUCT_NAMES_QUERY))
        for m in r.mappings():
            if m["sku"] and m["n"]:
                names[m["sku"]] = m["n"]
    except Exception:
        pass

    # ── przewidywany rozchód: średnia dzienna z 90 dni ───────
    # Okno 90-dniowe zamiast 30, bo wygładza pojedyncze duże zamówienia. Filtr statusów
    # ten sam, co w SALES_QUERY (per sklep), żeby liczby zgadzały się z resztą systemu.
    daily_sales: Dict[tuple, float] = {}
    if include_sales and horizon > 0:
        try:
            r = await db.execute(text(f"""
                SELECT UPPER(TRIM(oi.{settings.COL_ITEM_SKU})) AS sku,
                       LOWER(TRIM(o.shop)) AS shop,
                       SUM(oi.{settings.COL_ITEM_QTY}) AS qty
                FROM {settings.TABLE_ORDER_ITEMS} oi
                JOIN {settings.TABLE_ORDERS} o
                  ON o.{settings.COL_ORDER_ID} = oi.{settings.COL_ITEM_ORDER_ID} AND o.shop = oi.shop
                WHERE o.{settings.COL_ORDER_DATE} >= NOW() - INTERVAL '90 days'
                  {INCLUDED_STATUS_FILTER}
                  AND oi.{settings.COL_ITEM_SKU} IS NOT NULL AND TRIM(oi.{settings.COL_ITEM_SKU}) <> ''
                GROUP BY UPPER(TRIM(oi.{settings.COL_ITEM_SKU})), LOWER(TRIM(o.shop))
            """))
            for m in r.mappings():
                if m["sku"] and m["shop"]:
                    daily_sales[(m["sku"], m["shop"])] = float(m["qty"] or 0) / 90.0
        except Exception:
            daily_sales = {}

    # ── agregat per (SKU, magazyn) ───────────────────────────
    acc: Dict[tuple, dict] = {}

    def cell(sku_raw: str, firma: str) -> dict:
        key = ((sku_raw or "").strip().upper(), firma)
        if key not in acc:
            acc[key] = {"sku": (sku_raw or "").strip(), "nazwa": "", "firma_slug": firma,
                        "cbm": cbm_by_sku.get(key[0], 0.0), "stock_qty": 0, "incoming_qty": 0}
        return acc[key]

    # 1) stan fizyczny — magazyn GŁÓWNY (bez „w drodze" z ERP)
    for e in await build_stock_rows(db):
        firma = (e.get("firma_slug") or "amh").strip().lower()
        c = cell(e.get("sku") or "", firma)
        c["stock_qty"] += int(e.get("stan_glowny") or 0)
        if not c["nazwa"]:
            c["nazwa"] = e.get("nazwa") or ""

    # 2) w drodze — kontenery z datą wejścia na magazyn w horyzoncie
    timeline: Dict[tuple, dict] = {}
    for cont in await fetch_containers(db):
        if (cont.effective_status or cont.status) == "DELIVERED":
            continue
        arrival = getattr(cont, "warehouse_delivery_date", None)
        if not arrival or arrival <= today:
            # Bez daty nie wiemy, kiedy zajmie miejsce; data przeszła → towar
            # powinien już siedzieć w stanie głównym, doliczanie byłoby dublem.
            continue
        for it in (cont.items or []):
            key = (getattr(it, "sku", "") or "").strip().upper()
            if not key:
                continue
            qty = int(getattr(it, "quantity", 0) or 0)
            if qty <= 0:
                continue
            firma = firma_of.get(key, "amh")
            vol = cbm_by_sku.get(key, 0.0) * qty
            tkey = (arrival, cont.container_number or f"#{cont.id}")
            # PO obok numeru: skonsolidowany trzyma je na lotach, zwykły na sobie.
            # Front (containerLabel) podmienia numer roboczy „Draft-…" właśnie na to pole.
            po_c = (getattr(cont, "order_number", None) or "").strip()
            if not po_c:
                po_c = ", ".join(sorted({
                    (getattr(l, "order_number", None) or "").strip()
                    for l in (getattr(cont, "lots", None) or [])
                    if (getattr(l, "order_number", None) or "").strip()
                }))
            t = timeline.setdefault(tkey, {"date": arrival.isoformat(), "container_number": tkey[1],
                                           "order_number": po_c or None, "m3": 0.0, "firmy": {}})
            if scope in ("all", firma):
                t["m3"] += vol
                t["firmy"][firma] = round(t["firmy"].get(firma, 0.0) + vol, 3)
            if arrival <= cutoff:
                cell(getattr(it, "sku", key), firma)["incoming_qty"] += qty

    # ── wiersze + podsumowania per firma ─────────────────────
    per_firma: Dict[str, dict] = {}
    for slug, cap in caps.items():
        per_firma[slug] = {"slug": slug, "label": OCC_FIRMA_LABELS.get(slug, slug.upper()),
                           "capacity_m3": round(float(cap or 0), 2), "stock_m3": 0.0, "incoming_m3": 0.0,
                           "sold_m3": 0.0, "sku_count": 0, "over_count": 0, "no_cbm_count": 0}

    rows: List[dict] = []
    missing_sku, missing_units = 0, 0
    for (key, firma), c in acc.items():
        qty = c["stock_qty"] + c["incoming_qty"]
        if qty <= 0:
            continue
        no_cbm = c["cbm"] <= 0
        if no_cbm:
            missing_sku += 1
            missing_units += qty
        if firma not in per_firma:
            per_firma[firma] = {"slug": firma, "label": OCC_FIRMA_LABELS.get(firma, firma.upper()),
                                "capacity_m3": 0.0, "stock_m3": 0.0, "incoming_m3": 0.0, "sold_m3": 0.0,
                                "sku_count": 0, "over_count": 0, "no_cbm_count": 0}
        # Przewidywana sprzedaż zjada najpierw to, co stoi w hali, a dopiero potem to,
        # co dojedzie. Nie schodzimy poniżej zera — magazyn nie ma ujemnej objętości.
        sold_qty = min(qty, int(round(daily_sales.get((key, firma), 0.0) * horizon)))
        left_stock = max(0, c["stock_qty"] - sold_qty)
        left_incoming = max(0, qty - sold_qty - left_stock)

        stock_m3 = c["cbm"] * left_stock
        incoming_m3 = c["cbm"] * left_incoming
        sold_m3 = c["cbm"] * sold_qty
        per_firma[firma]["stock_m3"] += stock_m3
        per_firma[firma]["incoming_m3"] += incoming_m3
        per_firma[firma]["sold_m3"] += sold_m3
        # SKU bez kubatury zostaje w tabeli (żebyś wiedział, które uzupełnić), ale nie
        # udaje, że coś zajmuje — nie wchodzi do sumy m³ ani do licznika SKU firmy.
        if no_cbm:
            per_firma[firma]["no_cbm_count"] += 1
        else:
            per_firma[firma]["sku_count"] += 1
        rows.append({
            "sku": c["sku"], "nazwa": names.get(key) or c["nazwa"] or "", "firma_slug": firma,
            "no_cbm": no_cbm,
            "cbm_per_unit": round(c["cbm"], 3),
            "stock_qty": c["stock_qty"], "incoming_qty": c["incoming_qty"], "qty": qty,
            "sold_qty": sold_qty, "qty_left": qty - sold_qty,
            "stock_m3": round(stock_m3, 3), "incoming_m3": round(incoming_m3, 3),
            "sold_m3": round(sold_m3, 3),
            "volume_m3": round(stock_m3 + incoming_m3, 3),
        })

    # Udział produktu liczymy ZAWSZE względem hali, w której towar stoi — bo tam
    # fizycznie blokuje miejsce. Przy „Wszyscy" dokładamy udział w sumie hal.
    scope_cap = sum(f["capacity_m3"] for f in per_firma.values()) if scope == "all" else float(caps.get(scope, 0) or 0)
    for r_ in rows:
        firm_cap = per_firma[r_["firma_slug"]]["capacity_m3"]
        r_["share_firm_pct"] = round((r_["volume_m3"] / firm_cap) * 100, 2) if firm_cap > 0 else 0.0
        r_["share_scope_pct"] = round((r_["volume_m3"] / scope_cap) * 100, 2) if scope_cap > 0 else 0.0
        if r_.get("no_cbm"):
            r_["threshold_label"] = "Brak CBM"
            r_["threshold_tone"] = "pending"
            r_["over"] = False
            continue
        hit = _occ_match(thresholds["product"], r_["share_firm_pct"])
        r_["threshold_label"] = hit["label"]
        r_["threshold_tone"] = hit["tone"]
        r_["over"] = r_["share_firm_pct"] >= top_product["from_pct"]
        if r_["over"]:
            per_firma[r_["firma_slug"]]["over_count"] += 1

    if scope != "all":
        rows = [r_ for r_ in rows if r_["firma_slug"] == scope]

    for f in per_firma.values():
        used = f["stock_m3"] + f["incoming_m3"]
        f["stock_m3"] = round(f["stock_m3"], 2)
        f["incoming_m3"] = round(f["incoming_m3"], 2)
        f["sold_m3"] = round(f["sold_m3"], 2)
        f["used_m3"] = round(used, 2)
        f["free_m3"] = round(f["capacity_m3"] - used, 2)
        f["fill_pct"] = round((used / f["capacity_m3"]) * 100, 2) if f["capacity_m3"] > 0 else 0.0
        hit = _occ_match(thresholds["fill"], f["fill_pct"])
        f["threshold_label"] = hit["label"]
        f["threshold_tone"] = hit["tone"]

    firms = [per_firma[s] for s in sorted(per_firma, key=lambda x: (x != "amh", x))]
    in_scope = firms if scope == "all" else [f for f in firms if f["slug"] == scope]

    stock_m3 = round(sum(f["stock_m3"] for f in in_scope), 2)
    incoming_m3 = round(sum(f["incoming_m3"] for f in in_scope), 2)
    sold_m3 = round(sum(f["sold_m3"] for f in in_scope), 2)
    used_m3 = round(stock_m3 + incoming_m3, 2)
    capacity = round(sum(f["capacity_m3"] for f in in_scope), 2)
    fill_pct = round((used_m3 / capacity) * 100, 2) if capacity > 0 else 0.0
    fill_hit = _occ_match(thresholds["fill"], fill_pct)

    rows.sort(key=lambda x: x["volume_m3"], reverse=True)
    tl = sorted(timeline.values(), key=lambda t: t["date"])
    tl = [t for t in tl if t["m3"] > 0][:60]
    for t in tl:
        t["m3"] = round(t["m3"], 2)

    return {
        "scope": scope, "horizon_days": horizon,
        "as_of": today.isoformat(), "cutoff": cutoff.isoformat(),
        "capacity_m3": capacity,
        "stock_m3": stock_m3, "incoming_m3": incoming_m3, "sold_m3": sold_m3,
        "sales_included": bool(include_sales and horizon > 0), "sales_window_days": 90,
        "used_m3": used_m3,
        "free_m3": round(capacity - used_m3, 2),
        "fill_pct": fill_pct, "fill_label": fill_hit["label"], "fill_tone": fill_hit["tone"],
        "over_count": sum(1 for r_ in rows if r_["over"]),
        "over_threshold_pct": top_product["from_pct"], "over_threshold_label": top_product["label"],
        "firms": firms,
        "rows": rows,
        "timeline": tl,
        "missing_cbm": {"sku_count": sum(1 for r_ in rows if r_.get("no_cbm")),
                        "units": sum(r_["qty"] for r_ in rows if r_.get("no_cbm")),
                        "sku_count_all": missing_sku, "units_all": missing_units},
        "thresholds": thresholds,
        "caps": {k: round(float(v or 0), 2) for k, v in caps.items()},
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


@router.get("/reports/occupancy")
async def occupancy(
    scope: str = Query("all"), horizon: int = Query(0, ge=0, le=OCC_MAX_HORIZON),
    sales: bool = Query(True),
    db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_occupancy),
):
    """Zajętość hal w m³ na dziś + po dostawach do dnia `dziś + horizon`.

    `sales=false` wyłącza odejmowanie przewidywanego rozchodu (scenariusz „nic się nie sprzeda").
    """
    return await _occ_compute(db, resolve_scope(scope, user), horizon, include_sales=sales)


@router.get("/reports/occupancy/config")
async def occupancy_config(db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_occupancy)):
    """Pojemności hal i progi — do panelu ustawień raportu."""
    return {"caps": {k: round(float(v or 0), 2) for k, v in (await _occ_caps(db)).items()},
            "thresholds": await _occ_thresholds(db)}


@router.put("/reports/occupancy/config")
async def occupancy_config_save(
    payload: OccConfigIn, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_occupancy),
):
    """Zapis pojemności i progów. Progi nadpisujemy w całości — lista z UI jest źródłem prawdy."""
    for slug, val in (payload.caps or {}).items():
        s = (slug or "").strip().lower()
        if not s:
            continue
        if val is None or val < 0:
            raise HTTPException(400, f"Pojemność dla {s} nie może być ujemna")
        await db.execute(text("""
            INSERT INTO app_warehouse_capacity (firma_slug, capacity_m3, updated_at)
            VALUES (:slug, :cap, CURRENT_TIMESTAMP)
            ON CONFLICT (firma_slug) DO UPDATE SET capacity_m3 = EXCLUDED.capacity_m3, updated_at = CURRENT_TIMESTAMP
        """), {"slug": s, "cap": float(val)})

    for kind, items in (("product", payload.product), ("fill", payload.fill)):
        if items is None:
            continue
        clean = []
        for t in items:
            tone = t.tone if t.tone in OCC_TONES else "info"
            pct = max(0.0, min(float(t.from_pct or 0), 1000.0))
            label = (t.label or "").strip()
            if label:
                clean.append({"label": label[:40], "from_pct": pct, "tone": tone})
        if not clean:
            clean = _occ_defaults(kind)
        clean.sort(key=lambda x: x["from_pct"])
        # Pierwszy próg musi startować od 0 — inaczej wartości poniżej nie dostałyby etykiety.
        clean[0]["from_pct"] = 0.0
        await db.execute(text("DELETE FROM app_occupancy_thresholds WHERE kind = :k"), {"k": kind})
        for i, t in enumerate(clean):
            await db.execute(text("""
                INSERT INTO app_occupancy_thresholds (kind, label, from_pct, tone, sort_order, updated_at)
                VALUES (:k, :l, :p, :t, :o, CURRENT_TIMESTAMP)
            """), {"k": kind, "l": t["label"], "p": t["from_pct"], "t": t["tone"], "o": i})

    await db.commit()
    await log_audit(db, user, "OCCUPANCY_CONFIG_SAVED", "reports", "occupancy",
                    f"caps={payload.caps}")
    return {"caps": {k: round(float(v or 0), 2) for k, v in (await _occ_caps(db)).items()},
            "thresholds": await _occ_thresholds(db)}


# ── eksport XLSX ─────────────────────────────────────────────

def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, row_idx: int, headers: List[str]):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1F3864")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")


@router.get("/reports/kpi-range/xlsx")
async def kpi_range_xlsx(
    date_from: str = Query("", alias="from"), date_to: str = Query("", alias="to"),
    scope: str = Query("all"), group: str = Query("day"), slot: str = Query(""),
    fields: str = Query(""), live: bool = Query(False),
    db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    scope = resolve_scope(scope, user)
    data = (await _live_kpi(db, scope)) if live else (
        await kpi_range(date_from=date_from, date_to=date_to, scope=scope, group=group, slot=slot, db=db, user=user))
    chosen = [k.strip() for k in fields.split(",") if k.strip()] or [k for k, _ in KPI_FIELDS]
    cols = [(k, l) for k, l in KPI_FIELDS if k in chosen]

    wb = Workbook(); ws = wb.active; ws.title = "Raport zbiorczy"
    okres_x = "stan na teraz" if data.get("live") else f"{data['from']} … {data['to']}"
    ws["A1"] = f"Raport zbiorczy magazynu — {okres_x}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Zakres: {scope.upper()} · grupowanie: {'miesiąc' if group == 'month' else 'dzień'}"
    ws["A2"].font = Font(color="808080")

    head = ["Okres"] + [l for _, l in cols]
    _style_header(ws, 4, head)
    for j, row in enumerate(data["rows"], start=5):
        ws.cell(row=j, column=1, value=row["label"])
        for i, (k, _) in enumerate(cols, start=2):
            ws.cell(row=j, column=i, value=row[k]).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 16
    for i in range(len(cols)):
        ws.column_dimensions[chr(ord("B") + i)].width = 20
    return _xlsx_response(wb, f"raport_zbiorczy_{data['from']}_{data['to']}_{scope}.xlsx")


@router.get("/reports/sku/xlsx")
async def sku_xlsx(
    date_from: str = Query("", alias="from"), date_to: str = Query("", alias="to"),
    favorites_only: bool = Query(False), skus: str = Query(""), slot: str = Query(""),
    live: bool = Query(False), scope: str = Query("all"),
    db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(_require_reports),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    scope = resolve_scope(scope, user)
    data = (await _live_sku(db, favorites_only, skus, scope)) if live else (await sku_report(date_from=date_from, date_to=date_to, favorites_only=favorites_only,
                            skus=skus, slot=slot, scope=scope, db=db, user=user))
    rng = data["is_range"]

    wb = Workbook(); ws = wb.active; ws.title = "Magazyn per SKU"
    ws["A1"] = "Raport magazynu per SKU — stan na teraz" if data.get("live") else (f"Raport magazynu per SKU — {data['from']}" + (f" … {data['to']}" if rng else ""))
    ws["A1"].font = Font(bold=True, size=14)
    sub = []
    if favorites_only: sub.append("tylko obserwowane")
    if skus: sub.append("wybrane SKU")
    ws["A2"] = " · ".join(sub) if sub else "wszystkie SKU"
    ws["A2"].font = Font(color="808080")

    head = ["SKU", "Nazwa", "Firma", "Cena jedn.", "Magazyn główny", "W drodze", "W kontenerze", "Razem szt"]
    if rng:
        head += ["Szt. początek", "Szt. koniec", "Zmiana szt"]
    _style_header(ws, 4, head)

    for j, r in enumerate(data["rows"], start=5):
        vals = [r["sku"], r["nazwa"], r["firma_slug"], r["cena_jednostkowa"], r["stan_glowny"],
                r["stan_w_drodze"], r["w_kontenerze"], r["razem"]]
        if rng:
            vals += [r.get("razem_start"), r.get("razem_end"), r.get("delta_szt")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=j, column=i, value=v)
            if isinstance(v, float):
                c.number_format = "#,##0.00"

    t = data["totals"]
    last = 5 + len(data["rows"]) + 1
    ws.cell(row=last, column=1, value="RAZEM").font = Font(bold=True)
    for col_i, key in ((5, "units_glowny"), (6, "units_w_drodze"), (7, "units_kontener"), (8, "units")):
        ws.cell(row=last, column=col_i, value=t.get(key, 0)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 34
    for col in "CDEFGHI":
        ws.column_dimensions[col].width = 15
    name = f"raport_sku_{data['from']}" + (f"_{data['to']}" if rng else "") + ".xlsx"
    return _xlsx_response(wb, name)
