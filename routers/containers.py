"""Kontenery: CRUD, oznaczanie dostarczenia, załączniki (metadane), eksport do XLSX."""

import io
import asyncio
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import OperationalError, InterfaceError

from config import settings
from database import get_db, SessionLocal
from models import (
    ContainerStatus, ContainerOut, ContainerCreate, ContainerUpdate,
    AttachmentOut, AttachmentCreate, CurrentUser, SubiektWbiteIn,
)
from security import get_current_user, require_edit_containers, require_export, has_perm
from services.containers import fetch_containers, get_container_by_id

router = APIRouter(prefix="/api", tags=["containers"])


def _mask_container_financials(containers, user):
    """Serwerowe ukrycie cen w kontenerach dla usera bez viewFinancials."""
    if has_perm(user, "viewFinancials"):
        return containers
    for c in containers:
        c.total_value = 0.0
        c.koszt_transportu = None
        c.koszt_spedycji = None
        c.oplata_spedycji = None
        c.koszt_transportu_magazyn = None
        c.zaliczka_kwota = None
        c.balance_kwota = None
        for adv in c.advances:
            adv.kwota = None
        for it in c.items:
            it.unit_cost = None
        for lot in c.lots:
            lot.total_value = 0.0
            lot.zaliczka_kwota = None
            lot.balance_kwota = None
            for adv in lot.advances:
                adv.kwota = None
    return containers


def _advances_from(adv_list, z_proc, z_kwota, z_wal, z_data, default_cur="USD") -> List[dict]:
    """Normalizuje zaliczki do listy dict-ów. Puste wiersze (bez kwoty/daty/%) pomijane.
    Gdy front nie przysłał `advances`, spada na pojedynczą legacy zaliczkę (kompat ze starym
    frontem i z danymi sprzed migracji)."""
    out: List[dict] = []
    for a in (adv_list or []):
        if a.kwota is None and a.data is None and a.procent is None:
            continue
        out.append({"procent": a.procent, "kwota": a.kwota,
                    "waluta": (a.waluta or default_cur), "data": a.data})
    if not out and (z_kwota is not None or z_data is not None or z_proc is not None):
        out.append({"procent": z_proc, "kwota": z_kwota,
                    "waluta": (z_wal or default_cur), "data": z_data})
    return out


# ── Unikalność numerów + numer roboczy (Draft-<Producent>) ────────────────────
# Numer kontenera dostajemy dopiero po produkcji, a numer zamówienia/faktury od razu
# przy zamówieniu. Dlatego: container_number bywa pusty (nadajemy Draft-…), a oba
# numery muszą być unikalne — porównanie po UPPER(TRIM(...)), żeby „sk2605013 " i
# „SK2605013" były tym samym. PO żyje na kontenerze (nieskonsolidowany) ALBO na
# lotach (skonsolidowany), więc sprawdzamy obie tabele naraz.

DRAFT_PREFIX = "DRAFT-"


def _norm_nr(v: Optional[str]) -> Optional[str]:
    v = (v or "").strip()
    return v.upper() if v else None


def _is_draft(v: Optional[str]) -> bool:
    n = _norm_nr(v)
    return bool(n and n.startswith(DRAFT_PREFIX))


async def _next_draft_number(db: AsyncSession, *, manufacturer_id: Optional[int], consolidated: bool) -> str:
    """Kolejny wolny numer roboczy: Draft-Anji, Draft-Anji2, Draft-Anji3…

    Skonsolidowany kontener nie ma jednego dostawcy → Draft-Mix.
    """
    label = "Mix"
    if not consolidated and manufacturer_id:
        r = await db.execute(text(f"SELECT name FROM {settings.TABLE_MANUFACTURERS} WHERE id = :id"), {"id": manufacturer_id})
        nm = r.scalar()
        if nm and nm.strip():
            label = nm.strip().replace(" ", "-")
    base = f"Draft-{label}"

    r = await db.execute(
        text(f"""
            SELECT UPPER(TRIM(container_number))
            FROM {settings.TABLE_CONTAINERS}
            WHERE UPPER(TRIM(container_number)) LIKE :pat
        """),
        {"pat": base.upper() + "%"},
    )
    used = {row[0] for row in r if row[0]}
    if base.upper() not in used:
        return base
    n = 2
    while f"{base}{n}".upper() in used:
        n += 1
    return f"{base}{n}"


async def _assert_numbers_free(db: AsyncSession, *, container_number: Optional[str],
                               order_numbers: List[Optional[str]], exclude_cid: Optional[int] = None) -> None:
    """Rzuca 409, gdy numer kontenera albo numer zamówienia/faktury już istnieje.

    order_numbers = PO kontenera i/lub PO wszystkich lotów (zależnie od wariantu).
    """
    cn = _norm_nr(container_number)
    if cn:
        r = await db.execute(
            text(f"""
                SELECT container_number FROM {settings.TABLE_CONTAINERS}
                WHERE UPPER(TRIM(container_number)) = :v AND (:cid IS NULL OR id <> :cid)
                LIMIT 1
            """),
            {"v": cn, "cid": exclude_cid},
        )
        hit = r.scalar()
        if hit:
            raise HTTPException(409, f"Kontener o numerze „{hit}\u201d już istnieje")

    seen: set = set()
    for raw in order_numbers:
        v = _norm_nr(raw)
        if not v:
            continue
        if v in seen:
            raise HTTPException(409, f"Numer zamówienia „{raw}\u201d powtarza się w tym kontenerze")
        seen.add(v)

        r = await db.execute(
            text(f"""
                SELECT order_number FROM {settings.TABLE_CONTAINERS}
                WHERE UPPER(TRIM(order_number)) = :v AND (:cid IS NULL OR id <> :cid)
                LIMIT 1
            """),
            {"v": v, "cid": exclude_cid},
        )
        hit = r.scalar()
        if hit:
            raise HTTPException(409, f"Numer zamówienia „{hit}\u201d jest już użyty w innym kontenerze")

        r = await db.execute(
            text(f"""
                SELECT l.order_number FROM {settings.TABLE_CONTAINER_LOTS} l
                WHERE UPPER(TRIM(l.order_number)) = :v AND (:cid IS NULL OR l.container_id <> :cid)
                LIMIT 1
            """),
            {"v": v, "cid": exclude_cid},
        )
        hit = r.scalar()
        if hit:
            raise HTTPException(409, f"Numer zamówienia „{hit}\u201d jest już użyty w locie innego kontenera")


async def _insert_advances(db: AsyncSession, *, advances: List[dict],
                           container_id: Optional[int] = None, lot_id: Optional[int] = None) -> None:
    """Wstawia zaliczki podpięte pod kontener ALBO lot (dokładnie jedno z id)."""
    for pos, a in enumerate(advances):
        await db.execute(
            text(f"""
                INSERT INTO {settings.TABLE_CONTAINER_ADVANCES}
                (container_id, lot_id, position, procent, kwota, waluta, data)
                VALUES (:cid, :lid, :p, :proc, :kw, :wal, :dt)
            """),
            {"cid": container_id, "lid": lot_id, "p": pos,
             "proc": a["procent"], "kw": a["kwota"], "wal": a["waluta"], "dt": a["data"]},
        )


async def _replace_lots(db: AsyncSession, cid: int, lots) -> List[int]:
    """Usuwa loty kontenera i wstawia nowe (po kolei). Zwraca listę nowych id w kolejności.
    Zaliczki lotu lecą do app_container_advances (kaskada usuwa je przy DELETE lotu);
    1. zaliczkę mirror-ujemy do legacy zaliczka_* na locie (bezpieczny rollback)."""
    await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINER_LOTS} WHERE container_id = :c"), {"c": cid})
    ids: List[int] = []
    for pos, lot in enumerate(lots or []):
        default_cur = lot.waluta_towaru or "USD"
        advs = _advances_from(lot.advances, lot.zaliczka_procent, lot.zaliczka_kwota,
                              lot.zaliczka_waluta, lot.zaliczka_data, default_cur)
        first = advs[0] if advs else None
        rr = await db.execute(
            text(f"""
                INSERT INTO {settings.TABLE_CONTAINER_LOTS}
                (container_id, manufacturer_id, order_number, position,
                 waluta_towaru, zaliczka_procent, zaliczka_kwota, zaliczka_waluta, zaliczka_data,
                 balance_kwota, balance_waluta, zaplacono_data)
                VALUES (:c, :m, :o, :p, :wal, :zp, :zk, :zwal, :zd, :bal, :bwal, :pd)
                RETURNING id
            """),
            {"c": cid, "m": lot.manufacturer_id, "o": (lot.order_number or None), "p": pos,
             "wal": default_cur,
             "zp": (first["procent"] if first else None),
             "zk": (first["kwota"] if first else None),
             "zwal": (first["waluta"] if first else default_cur),
             "zd": (first["data"] if first else None),
             "bal": lot.balance_kwota, "bwal": (lot.balance_waluta or default_cur),
             "pd": lot.zaplacono_data},
        )
        lid = rr.scalar_one()
        await _insert_advances(db, advances=advs, lot_id=lid)
        ids.append(lid)
    return ids


def _resolve_lot(lot_ref: Optional[int], lot_ids: List[int]) -> Optional[int]:
    if lot_ref is None:
        return None
    if 0 <= lot_ref < len(lot_ids):
        return lot_ids[lot_ref]
    return None


@router.get("/containers/export/csv")
async def export_containers_xlsx(db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_export)):
    """Eksport kontenerów do Excela (XLSX)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    containers = await fetch_containers(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontenery"

    headers = [
        "Nr kontenera", "Nr zamówienia", "Producent", "Typ", "Status",
        "Data zamówienia", "ETA", "SKU", "Nazwa produktu",
        "Ilość", "Cena jednostkowa", "Wartość", "CBM total",
        "Folder", "Subiekt", "Koszt transportu", "Koszt spedycji", "Opłata spedycji", "Transport do magazynu (PLN)",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1c1917", end_color="1c1917", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_label = {"ORDERED": "Zamówione", "IN_PRODUCTION": "W produkcji", "IN_TRANSIT": "W drodze", "CUSTOMS": "Odprawa celna", "DELIVERED": "Dostarczone"}

    for c in containers:
        lot_map = {l.id: (l.order_number, l.manufacturer_name) for l in c.lots}
        for it in c.items:
            cena = float(it.unit_cost) if it.unit_cost else 0
            wartosc = cena * it.quantity
            po, mfr = c.order_number, c.manufacturer_name
            if c.is_consolidated and it.lot_id in lot_map:
                po, mfr = lot_map[it.lot_id]
            ws.append([
                c.container_number, po or "",
                mfr or "", c.container_type_name or "",
                status_label.get(c.effective_status, c.effective_status),
                c.order_date.isoformat(), c.eta_date.isoformat(),
                it.sku, it.product_name or "",
                it.quantity, cena, wartosc, it.total_cbm,
                c.folder or "", c.subiekt_nr or "",
                (c.koszt_transportu if c.koszt_transportu is not None else ""),
                (c.koszt_spedycji if c.koszt_spedycji is not None else ""),
                (c.oplata_spedycji if c.oplata_spedycji is not None else ""),
                (c.koszt_transportu_magazyn if c.koszt_transportu_magazyn is not None else ""),
            ])

    column_widths = [16, 16, 18, 8, 14, 14, 14, 12, 35, 8, 14, 14, 10, 10, 12, 16, 15, 15, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"kontenery_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/containers", response_model=List[ContainerOut])
async def list_containers(status: Optional[ContainerStatus] = None, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    return _mask_container_financials(await fetch_containers(db, status), user)


@router.get("/containers/{cid}", response_model=ContainerOut)
async def get_container(cid: int, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    c = await get_container_by_id(db, cid)
    _mask_container_financials([c], user)
    return c


@router.post("/containers", response_model=ContainerOut, status_code=201)
async def create_container(payload: ContainerCreate, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    if payload.eta_date < payload.order_date:
        raise HTTPException(400, "ETA nie może być przed datą zamówienia")

    cons = payload.is_consolidated

    # Numer kontenera bywa nieznany przy zamówieniu (dostajemy go po produkcji) →
    # nadajemy roboczy Draft-<Producent>. Numery pilnujemy przed INSERT-em, żeby
    # użytkownik dostał czytelny 409 zamiast surowego błędu unikalnego indeksu.
    nr = (payload.container_number or "").strip()
    if not nr:
        nr = await _next_draft_number(db, manufacturer_id=payload.manufacturer_id, consolidated=cons)
    order_nums = ([lot.order_number for lot in (payload.lots or [])] if cons else [payload.order_number])
    await _assert_numbers_free(db, container_number=nr, order_numbers=order_nums)

    default_cur = payload.waluta_towaru or "USD"
    # Zaliczki kontenera tylko dla wariantu nieskonsolidowanego (przy konsolidacji siedzą w lotach).
    cont_advs = [] if cons else _advances_from(payload.advances, payload.zaliczka_procent,
                                               payload.zaliczka_kwota, payload.zaliczka_waluta,
                                               payload.zaliczka_data, default_cur)
    first = cont_advs[0] if cont_advs else None
    r = await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_CONTAINERS}
            (container_number, order_number, container_type_id, manufacturer_id, order_date, eta_date, status, notes, is_consolidated,
             koszt_transportu, koszt_spedycji, koszt_transportu_magazyn, folder, subiekt_nr,
             waluta_towaru, zaliczka_procent, zaliczka_kwota, zaliczka_waluta, zaliczka_data,
             balance_kwota, balance_waluta, zaplacono_data, expected_delivery_date)
            VALUES (:n, :on, :tid, :mid, :od, :eta, :st, :no, :cons,
                    :kt, :ks, :ktm, :fol, :sub,
                    :wal, :zp, :zk, :zwal, :zd, :bal, :bwal, :pd, :edd)
            RETURNING id
        """),
        {"n": nr,
         "on": (None if cons else payload.order_number),
         "tid": payload.container_type_id,
         "mid": (None if cons else payload.manufacturer_id),
         "od": payload.order_date, "eta": payload.eta_date,
         "st": payload.status, "no": payload.notes, "cons": cons,
         "kt": payload.koszt_transportu, "ks": payload.koszt_spedycji,
         "ktm": payload.koszt_transportu_magazyn,   # PLN — zawsze na kontenerze
         "fol": (payload.folder or None), "sub": (payload.subiekt_nr or None),
         # legacy zaliczka_* = 1. zaliczka z listy (mirror dla rollbacku); przy konsolidacji NULL
         "wal": (None if cons else default_cur),
         "zp": (first["procent"] if first else None),
         "zk": (first["kwota"] if first else None),
         "zwal": (None if cons else (first["waluta"] if first else default_cur)),
         "zd": (first["data"] if first else None),
         "bal": (None if cons else payload.balance_kwota),
         "bwal": (None if cons else (payload.balance_waluta or default_cur)),
         "pd": (None if cons else payload.zaplacono_data),
         "edd": payload.expected_delivery_date}
    )
    cid = r.scalar_one()

    if not cons:
        await _insert_advances(db, advances=cont_advs, container_id=cid)

    lot_ids = await _replace_lots(db, cid, payload.lots) if payload.is_consolidated else []

    for item in payload.items:
        lid = _resolve_lot(item.lot_ref, lot_ids) if payload.is_consolidated else None
        await db.execute(
            text(f"INSERT INTO {settings.TABLE_CONTAINER_ITEMS} (container_id, sku, quantity, unit_cost, lot_id) VALUES (:c, :s, :q, :u, :l)"),
            {"c": cid, "s": item.sku, "q": item.quantity, "u": item.unit_cost, "l": lid}
        )

    await db.commit()
    return await get_container_by_id(db, cid)


@router.patch("/containers/{cid}", response_model=ContainerOut)
async def update_container(cid: int, payload: ContainerUpdate, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    cur = (await db.execute(
        text(f"SELECT container_number, status, is_consolidated FROM {settings.TABLE_CONTAINERS} WHERE id = :id"),
        {"id": cid},
    )).mappings().first()
    if not cur:
        raise HTTPException(404)

    cons_now = payload.is_consolidated if payload.is_consolidated is not None else bool(cur["is_consolidated"])

    # Wyczyszczenie numeru w formularzu nie może zostawić pustego pola — wracamy do Draft-…
    if "container_number" in payload.model_fields_set and not (payload.container_number or "").strip():
        payload.container_number = await _next_draft_number(
            db,
            manufacturer_id=(payload.manufacturer_id if payload.manufacturer_id is not None else None),
            consolidated=cons_now,
        )

    # Numery muszą zostać unikalne także po edycji (siebie samego pomijamy).
    order_nums: List[Optional[str]] = []
    if payload.lots is not None:
        order_nums = [lot.order_number for lot in payload.lots]
    elif not cons_now and "order_number" in payload.model_fields_set:
        order_nums = [payload.order_number]
    await _assert_numbers_free(
        db,
        container_number=(payload.container_number if "container_number" in payload.model_fields_set else None),
        order_numbers=order_nums,
        exclude_cid=cid,
    )

    updates = []
    params = {"id": cid}
    for field in ["container_number", "container_type_id", "order_date", "eta_date", "status", "notes"]:
        v = getattr(payload, field)
        if v is not None:
            updates.append(f"{field} = :{field}")
            params[field] = v

    cons = payload.is_consolidated
    if cons is None:
        # częściowa aktualizacja bez informacji o konsolidacji — stare zachowanie
        for field in ["manufacturer_id", "order_number"]:
            v = getattr(payload, field)
            if v is not None:
                updates.append(f"{field} = :{field}")
                params[field] = v
    else:
        updates.append("is_consolidated = :cons")
        params["cons"] = cons
        if cons:
            updates.append("manufacturer_id = NULL")
            updates.append("order_number = NULL")
        else:
            updates.append("manufacturer_id = :mid")
            params["mid"] = payload.manufacturer_id
            updates.append("order_number = :onum")
            params["onum"] = payload.order_number

    # Koszty spedycji + dokumenty — zawsze na kontenerze; ruszamy tylko pola faktycznie
    # przysłane (model_fields_set), żeby null z formularza mógł je wyczyścić.
    fset = payload.model_fields_set
    if "koszt_transportu" in fset:
        updates.append("koszt_transportu = :kt"); params["kt"] = payload.koszt_transportu
    if "koszt_spedycji" in fset:
        updates.append("koszt_spedycji = :ks"); params["ks"] = payload.koszt_spedycji
    if "koszt_transportu_magazyn" in fset:
        updates.append("koszt_transportu_magazyn = :ktm"); params["ktm"] = payload.koszt_transportu_magazyn
    if "folder" in fset:
        updates.append("folder = :fol"); params["fol"] = (payload.folder or None)
    if "subiekt_nr" in fset:
        updates.append("subiekt_nr = :sub"); params["sub"] = (payload.subiekt_nr or None)

    # Płatności na kontenerze: przy konsolidacji przenoszą się do lotów → czyścimy;
    # w wariancie nieskonsolidowanym — waluta/balance z payloadu (sterowane fset).
    default_cur = payload.waluta_towaru or "USD"
    simple_pay = [
        ("waluta_towaru", "wal", default_cur),
        ("balance_kwota", "bal", payload.balance_kwota),
        ("balance_waluta", "bwal", (payload.balance_waluta or default_cur)),
        ("zaplacono_data", "pd", payload.zaplacono_data),
    ]
    # Zaliczki (rata) na poziomie kontenera — tylko wariant nieskonsolidowany.
    #   cons True         → czyścimy (dane siedzą w lotach),
    #   cons False/None   → przebudowa gdy front przysłał `advances` albo dotknął legacy zaliczka_*.
    # container_advs: None = nie ruszaj tabeli zaliczek; [] = wyczyść; [.] = zastąp.
    adv_touched = ("advances" in fset) or any(
        k in fset for k in ("zaliczka_procent", "zaliczka_kwota", "zaliczka_waluta", "zaliczka_data"))
    container_advs: Optional[List[dict]] = None
    if cons is True:
        for col, _ph, _val in simple_pay:
            updates.append(f"{col} = NULL")
        for col in ("zaliczka_procent", "zaliczka_kwota", "zaliczka_waluta", "zaliczka_data"):
            updates.append(f"{col} = NULL")
        container_advs = []
    else:
        for col, ph, val in simple_pay:
            if col in fset:
                updates.append(f"{col} = :{ph}"); params[ph] = val
        if adv_touched:
            container_advs = _advances_from(payload.advances, payload.zaliczka_procent,
                                            payload.zaliczka_kwota, payload.zaliczka_waluta,
                                            payload.zaliczka_data, default_cur)
            first = container_advs[0] if container_advs else None
            updates.append("zaliczka_procent = :zp"); params["zp"] = (first["procent"] if first else None)
            updates.append("zaliczka_kwota = :zk");   params["zk"] = (first["kwota"] if first else None)
            updates.append("zaliczka_waluta = :zwal"); params["zwal"] = (first["waluta"] if first else default_cur)
            updates.append("zaliczka_data = :zd");    params["zd"] = (first["data"] if first else None)

    # Data dostawy na magazyn (delivered_date):
    #   • gdy front przysyła delivered_date wprost → to źródło prawdy; ręczny wpis DOMYKA
    #     status (blokuje na DELIVERED), żeby kontener wypadł z „aktywnych" bez czekania na
    #     auto-dostawę z ETA (scenariusz: ktoś był na urlopie i klika datę z ręki później).
    #     Wyczyszczenie (null) zdejmuje ręczną datę → KPI wraca do auto (ETA + odprawa).
    #   • w przeciwnym razie datą steruje zmiana statusu (jak dotąd).
    if "delivered_date" in fset:
        if payload.delivered_date is not None:
            updates.append("delivered_date = :dd"); params["dd"] = payload.delivered_date
            if payload.status is None:            # brak jawnego statusu → domknij na DELIVERED
                updates.append("status = 'DELIVERED'")
        else:
            updates.append("delivered_date = NULL")
    elif payload.status is not None:
        if payload.status == "DELIVERED":
            updates.append("delivered_date = COALESCE(delivered_date, CURRENT_DATE)")
        else:
            updates.append("delivered_date = NULL")

    # Wpisanie prawdziwego numeru w miejsce roboczego Draft-… znaczy, że produkcja się
    # skończyła i kontener ruszył → przestawiamy status na „W drodze". Tylko gdy front nie
    # narzucił statusu jawnie i gdy kontener stoi jeszcze na wcześniejszym etapie.
    if ("container_number" in fset and payload.container_number
            and _is_draft(cur["container_number"]) and not _is_draft(payload.container_number)
            and payload.status is None and cur["status"] in ("ORDERED", "IN_PRODUCTION")):
        updates.append("status = 'IN_TRANSIT'")

    # Spodziewana dostawa („u nas"): umówiona data odbioru, znana zwykle w trakcie odprawy.
    # Świadomie NIE dotyka statusu — data z przyszłości ma zostawić kontener w „Odprawie
    # celnej" (z licznikiem do tej daty), a nie zapalić „Dostarczono". Domyka dopiero
    # delivered_date. Wysłanie null czyści → KPI wraca do szacunku ETA + okno odprawy.
    if "expected_delivery_date" in fset:
        if payload.expected_delivery_date is not None:
            updates.append("expected_delivery_date = :edd"); params["edd"] = payload.expected_delivery_date
        else:
            updates.append("expected_delivery_date = NULL")

    if updates:
        updates.append("updated_at = CURRENT_TIMESTAMP")
        await db.execute(text(f"UPDATE {settings.TABLE_CONTAINERS} SET {', '.join(updates)} WHERE id = :id"), params)

    # Zaliczki kontenera w osobnej tabeli: przebuduj tylko gdy front je ruszył (albo konsolidacja czyści).
    if container_advs is not None:
        await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINER_ADVANCES} WHERE container_id = :c"), {"c": cid})
        await _insert_advances(db, advances=container_advs, container_id=cid)

    if payload.items is not None:
        await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINER_ITEMS} WHERE container_id = :cid"), {"cid": cid})
        # cons=True → wstaw przysłane loty; cons=False → wyczyść loty (sieroty nie zostają);
        # cons=None (częściowa aktualizacja) → ruszamy loty tylko gdy front je przysłał.
        use_lots = bool(cons) if cons is not None else (payload.lots is not None)
        rebuild = (cons is not None) or (payload.lots is not None)
        lot_ids = await _replace_lots(db, cid, payload.lots if use_lots else []) if rebuild else []
        for item in payload.items:
            lid = _resolve_lot(item.lot_ref, lot_ids) if use_lots else None
            await db.execute(
                text(f"INSERT INTO {settings.TABLE_CONTAINER_ITEMS} (container_id, sku, quantity, unit_cost, lot_id) VALUES (:c, :s, :q, :u, :l)"),
                {"c": cid, "s": item.sku, "q": item.quantity, "u": item.unit_cost, "l": lid}
            )
    elif payload.lots is not None:
        await _replace_lots(db, cid, payload.lots)

    await db.commit()
    return await get_container_by_id(db, cid)


@router.delete("/containers/{cid}", status_code=204)
async def delete_container(cid: int, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_CONTAINERS} WHERE id = :id"), {"id": cid})
    await db.commit()
    if r.rowcount == 0:
        raise HTTPException(404)


@router.post("/containers/{cid}/deliver", response_model=ContainerOut)
async def deliver_container(cid: int, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    await db.execute(text(f"UPDATE {settings.TABLE_CONTAINERS} SET status = 'DELIVERED', delivered_date = COALESCE(delivered_date, CURRENT_DATE), updated_at = CURRENT_TIMESTAMP WHERE id = :id"), {"id": cid})
    await db.commit()
    return await get_container_by_id(db, cid)


@router.post("/containers/{cid}/subiekt-wbite", response_model=ContainerOut)
async def set_subiekt_wbite(cid: int, payload: SubiektWbiteIn, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    """Kropka „dodano do Subiektu": zielona = towar wbity do magazynu „w drodze" w Subiekcie
    (wypada z „Kontenerów w drodze", liczony z magazynu subiektowego). lot_id=None → kontener."""
    at = "CURRENT_DATE" if payload.value else "NULL"
    if payload.lot_id is not None:
        r = await db.execute(
            text(f"UPDATE {settings.TABLE_CONTAINER_LOTS} SET subiekt_wbite = :v, subiekt_wbite_at = {at} WHERE id = :lid AND container_id = :cid"),
            {"v": payload.value, "lid": payload.lot_id, "cid": cid},
        )
        if r.rowcount == 0:
            raise HTTPException(404, "Lot nie należy do tego kontenera")
    else:
        await db.execute(
            text(f"UPDATE {settings.TABLE_CONTAINERS} SET subiekt_wbite = :v, subiekt_wbite_at = {at}, updated_at = CURRENT_TIMESTAMP WHERE id = :cid"),
            {"v": payload.value, "cid": cid},
        )
    await db.commit()
    return await get_container_by_id(db, cid)


MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024  # 10 MB


def _human_size(n: int) -> str:
    if n >= 1024 * 1024:
        return f"{n / (1024 * 1024):.1f} MB"
    if n >= 1024:
        return f"{n / 1024:.0f} KB"
    return f"{n} B"


def _guess_type(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "pdf":
        return "pdf"
    if ext in ("xlsx", "xls"):
        return "excel"
    if ext in ("png", "jpg", "jpeg", "webp", "gif"):
        return "image"
    return ext or "other"


@router.post("/containers/{cid}/attachments", response_model=AttachmentOut, status_code=201)
async def add_attachment(cid: int, file: UploadFile = File(...), user: CurrentUser = Depends(require_edit_containers)):
    """Wgrywa plik (zawartość w bazie jako BYTEA). Ponawia zapis na świeżym
    połączeniu przy chwilowych błędach poolера Supabase (działa za 1. razem)."""
    data = await file.read()
    if not data:
        raise HTTPException(400, "Pusty plik")
    if len(data) > MAX_ATTACHMENT_BYTES:
        raise HTTPException(413, "Plik za duży (max 10 MB)")
    fname = file.filename or "plik"
    ftype = _guess_type(fname)
    fsize = _human_size(len(data))
    ctype = file.content_type or "application/octet-stream"
    params = {"c": cid, "n": fname, "t": ftype, "s": fsize, "ct": ctype, "d": data}
    sql = text(f"""
        INSERT INTO {settings.TABLE_ATTACHMENTS} (container_id, filename, file_type, file_size, content_type, file_data)
        VALUES (:c, :n, :t, :s, :ct, :d) RETURNING id, uploaded_at
    """)
    last_err: Exception | None = None
    for attempt in range(4):
        try:
            async with SessionLocal() as db:
                r = await db.execute(sql, params)
                row = r.first()
                await db.commit()
                return AttachmentOut(id=row.id, filename=fname, file_type=ftype, file_size=fsize, uploaded_at=row.uploaded_at)
        except (OperationalError, InterfaceError) as e:
            last_err = e
            await asyncio.sleep(0.4 * (attempt + 1))
    raise HTTPException(503, f"Zapis załącznika nieudany po kilku próbach: {last_err}")


@router.get("/attachments/{aid}/download")
async def download_attachment(aid: int, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    """Zwraca zawartość pliku załącznika."""
    r = await db.execute(text(f"SELECT filename, content_type, file_data FROM {settings.TABLE_ATTACHMENTS} WHERE id = :id"), {"id": aid})
    row = r.first()
    if not row or row.file_data is None:
        raise HTTPException(404, "Plik nie znaleziony (mógł być dodany przed włączeniem przechowywania)")
    data = bytes(row.file_data)
    ctype = row.content_type or "application/octet-stream"
    return StreamingResponse(io.BytesIO(data), media_type=ctype, headers={"Content-Disposition": f'attachment; filename="{row.filename}"'})


@router.delete("/attachments/{aid}", status_code=204)
async def delete_attachment(aid: int, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(require_edit_containers)):
    r = await db.execute(text(f"DELETE FROM {settings.TABLE_ATTACHMENTS} WHERE id = :id"), {"id": aid})
    await db.commit()
    if r.rowcount == 0:
        raise HTTPException(404)
