"""Produkty: lista z prognozą, edycja lead-time/atrybutów, projekcja stanu,
import atrybutów, eksport do XLSX, ulubione."""

import io
from datetime import date, timedelta
from typing import List

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from config import settings
from database import get_db
from models import (
    ProductSummary, LeadTimeUpdate, ProductAttrsUpdate,
    StockProjectionPoint, ImportRow, ImportResult, CurrentUser,
)
from security import get_current_user, has_perm
from services.products import fetch_products, get_product

router = APIRouter(prefix="/api", tags=["products"])


def _mask_financials(products, user):
    """Serwerowe ukrycie cen: zeruje pola finansowe dla usera bez viewFinancials.
    Front i tak maskuje wizualnie — to zamyka wyciek wartości w payloadzie (zakładka Network)."""
    if has_perm(user, "viewFinancials"):
        return products
    for p in products:
        p.stock_value = 0.0
        p.purchase_price = 0.0
        p.cena_zakupu_manual = None
    return products


@router.get("/products", response_model=List[ProductSummary])
async def list_products(include: str = Query("ACTIVE,ACTIVE_NO_STOCK"), shop: str = Query(""), db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    # shop="" = wszystkie sklepy (suma); "amh"/"acti"/"veluxa" = sprzedaż i stan tylko danego sklepu (Faza 3).
    allowed = set(s.strip().upper() for s in include.split(",") if s.strip())
    return _mask_financials(await fetch_products(db, allowed, shop), user)


@router.get("/products/{sku}", response_model=ProductSummary)
async def get_product_endpoint(sku: str, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    p = await get_product(db, sku)
    _mask_financials([p], user)
    return p


@router.put("/products/{sku:path}/lead-time", response_model=ProductSummary)
async def update_lead_time(sku: str, payload: LeadTimeUpdate, db: AsyncSession = Depends(get_db)):
    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_LEAD_TIMES} (sku, lead_time_days, updated_at)
            VALUES (:sku, :lt, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET lead_time_days = EXCLUDED.lead_time_days, updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "lt": payload.lead_time_days}
    )
    await db.commit()
    return await get_product(db, sku)


@router.put("/products/{sku:path}/attrs", response_model=ProductSummary)
async def update_attrs(sku: str, payload: ProductAttrsUpdate, db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    existing = await db.execute(text(f"SELECT cbm_per_unit, manufacturer_id, firma_id, seasonality_enabled, ean, forced_status, cena_zakupu FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": sku})
    e = existing.first()
    cbm = payload.cbm_per_unit if payload.cbm_per_unit is not None else (float(e.cbm_per_unit) if e else 0)
    # manufacturer_id: 0 = odepnij producenta; None = nie zmieniaj; >0 = ustaw
    if payload.manufacturer_id is not None:
        mfr = None if payload.manufacturer_id == 0 else payload.manufacturer_id
    else:
        mfr = e.manufacturer_id if e else None
    # firma_id: 0 = odepnij (→ domyślnie AMH); None = nie zmieniaj; >0 = ustaw
    if payload.firma_id is not None:
        firma = None if payload.firma_id == 0 else payload.firma_id
    else:
        firma = e.firma_id if e else None
    seas = payload.seasonality_enabled if payload.seasonality_enabled is not None else (e.seasonality_enabled if e else False)
    ean = payload.ean if payload.ean is not None else (e.ean if e else None)
    if ean is not None and not ean.strip():
        ean = None

    # Forced status: "AUTO" lub "" lub null = wyczyść
    forced = payload.forced_status if payload.forced_status is not None else (e.forced_status if e else None)
    if forced in ("", "AUTO", "auto", None):
        forced = None
    elif forced not in ("ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"):
        raise HTTPException(400, f"Niepoprawny status: {forced}. Dozwolone: ACTIVE, ACTIVE_NO_STOCK, DEAD_STOCK, INACTIVE, AUTO")

    # Cena zakupu (ręczna): None = nie zmieniaj; <=0 = wyczyść override; >0 = ustaw.
    # Dane finansowe — zapis TYLKO z uprawnieniem viewFinancials (guard po stronie serwera).
    if payload.cena_zakupu is not None:
        if not has_perm(user, "viewFinancials"):
            raise HTTPException(403, "Brak uprawnień do edycji ceny zakupu (viewFinancials)")
        cena = None if payload.cena_zakupu <= 0 else round(float(payload.cena_zakupu), 2)
    else:
        cena = (float(e.cena_zakupu) if (e and e.cena_zakupu is not None) else None)

    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, cbm_per_unit, manufacturer_id, firma_id, seasonality_enabled, ean, forced_status, cena_zakupu, updated_at)
            VALUES (:sku, :cbm, :mfr, :firma, :seas, :ean, :forced, :cena, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET
                cbm_per_unit = EXCLUDED.cbm_per_unit,
                manufacturer_id = EXCLUDED.manufacturer_id,
                firma_id = EXCLUDED.firma_id,
                seasonality_enabled = EXCLUDED.seasonality_enabled,
                ean = EXCLUDED.ean,
                forced_status = EXCLUDED.forced_status,
                cena_zakupu = EXCLUDED.cena_zakupu,
                updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "cbm": cbm, "mfr": mfr, "firma": firma, "seas": seas, "ean": ean, "forced": forced, "cena": cena}
    )
    await db.commit()
    return _mask_financials([await get_product(db, sku)], user)[0]


@router.get("/products/{sku:path}/projection", response_model=List[StockProjectionPoint])
async def projection(sku: str, days: int = 180, db: AsyncSession = Depends(get_db)):
    product = await get_product(db, sku)
    today = date.today()
    base_daily = product.avg_monthly_weighted / 30
    eta_map = {d.eta_date: d.quantity for d in product.incoming_deliveries}
    eta_names = {d.eta_date: f"#{d.container_number} +{d.quantity}" for d in product.incoming_deliveries}
    points = []
    current = float(product.stock)
    for offset in range(0, days + 1):
        cd = today + timedelta(days=offset)
        ev = None
        if cd in eta_map:
            current += eta_map[cd]
            ev = eta_names[cd]
        if offset > 0:
            current -= base_daily
        points.append(StockProjectionPoint(date=cd, stock=max(0, int(current)), event=ev))
    return points


@router.post("/products/import", response_model=ImportResult)
async def import_products(rows: List[ImportRow], db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    """Import atrybutów dla produktów - z UI lub bezpośrednio JSON-em."""
    can_fin = has_perm(user, "viewFinancials")  # cena zakupu tylko dla uprawnionych
    products_result = await db.execute(text(f"SELECT {settings.COL_PRODUCT_SKU} as sku FROM {settings.TABLE_PRODUCTS}"))
    valid_skus = {r._mapping["sku"].strip().lower(): r._mapping["sku"] for r in products_result}

    mfr_result = await db.execute(text(f"SELECT id, name FROM {settings.TABLE_MANUFACTURERS}"))
    mfr_map = {r._mapping["name"].strip().lower(): r._mapping["id"] for r in mfr_result}

    updated = 0
    skipped = 0
    errors = []

    for row in rows:
        sku_key = row.sku.strip().lower()
        if sku_key not in valid_skus:
            skipped += 1
            errors.append(f"{row.sku}: nie znaleziono w bazie")
            continue
        real_sku = valid_skus[sku_key]

        mfr_id = None
        if row.manufacturer_name:
            mfr_id = mfr_map.get(row.manufacturer_name.strip().lower())
            if mfr_id is None and row.manufacturer_name.strip():
                r = await db.execute(
                    text(f"INSERT INTO {settings.TABLE_MANUFACTURERS} (name, color) VALUES (:n, :c) ON CONFLICT (name) DO UPDATE SET name=EXCLUDED.name RETURNING id"),
                    {"n": row.manufacturer_name.strip(), "c": "#6b7280"}
                )
                mfr_id = r.scalar_one()
                mfr_map[row.manufacturer_name.strip().lower()] = mfr_id

        try:
            existing = await db.execute(text(f"SELECT cbm_per_unit, manufacturer_id, seasonality_enabled, cena_zakupu FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": real_sku})
            e = existing.first()
            cbm = row.cbm if row.cbm is not None else (float(e.cbm_per_unit) if e else 0)
            new_mfr = mfr_id if mfr_id is not None else (e.manufacturer_id if e else None)
            new_seas = row.seasonality_enabled if row.seasonality_enabled is not None else (e.seasonality_enabled if e else False)
            # Cena zakupu: tylko z uprawnieniem; puste = zostaw; <=0 = wyczyść; >0 = ustaw
            prev_cena = float(e.cena_zakupu) if (e and e.cena_zakupu is not None) else None
            if can_fin and row.cena_zakupu is not None:
                new_cena = None if row.cena_zakupu <= 0 else round(float(row.cena_zakupu), 2)
            else:
                new_cena = prev_cena

            await db.execute(text(f"""
                INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, cbm_per_unit, manufacturer_id, seasonality_enabled, cena_zakupu, updated_at)
                VALUES (:sku, :cbm, :mfr, :seas, :cena, CURRENT_TIMESTAMP)
                ON CONFLICT (sku) DO UPDATE SET cbm_per_unit=EXCLUDED.cbm_per_unit, manufacturer_id=EXCLUDED.manufacturer_id, seasonality_enabled=EXCLUDED.seasonality_enabled, cena_zakupu=EXCLUDED.cena_zakupu, updated_at=CURRENT_TIMESTAMP
            """), {"sku": real_sku, "cbm": cbm, "mfr": new_mfr, "seas": new_seas, "cena": new_cena})

            if row.lead_time_days is not None and 1 <= row.lead_time_days <= 365:
                await db.execute(text(f"""
                    INSERT INTO {settings.TABLE_LEAD_TIMES} (sku, lead_time_days, updated_at)
                    VALUES (:sku, :lt, CURRENT_TIMESTAMP)
                    ON CONFLICT (sku) DO UPDATE SET lead_time_days=EXCLUDED.lead_time_days, updated_at=CURRENT_TIMESTAMP
                """), {"sku": real_sku, "lt": row.lead_time_days})

            updated += 1
        except Exception as ex:
            errors.append(f"{row.sku}: {str(ex)[:100]}")
            skipped += 1

    await db.commit()
    return ImportResult(total=len(rows), updated=updated, skipped=skipped, errors=errors[:20])


@router.get("/products/export/csv")
async def export_xlsx(include: str = Query("ACTIVE,ACTIVE_NO_STOCK"), favorites_only: bool = Query(False), shop: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Eksport produktów do Excela (XLSX) - polskie znaki zawsze działają.
    shop="" = wszystkie sklepy; "amh"/"acti"/"veluxa" = liczby danego sklepu (zgodnie z wybraną zakładką)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    allowed = set(s.strip().upper() for s in include.split(",") if s.strip())
    products = await fetch_products(db, allowed, shop)

    if favorites_only:
        products = [p for p in products if p.is_favorite]

    wb = Workbook()
    ws = wb.active
    ws.title = "Produkty"

    headers = [
        "SKU", "Nazwa", "EAN", "Producent", "Stan", "Cena zakupu", "Wartość PLN",
        "W drodze", "CBM", "Lead time (dni)",
        "Sprzedaż 1m", "Sprzedaż 2m", "Sprzedaż 3m", "Sprzedaż 4m",
        "YoY 30d", "YoY +30d", "Średnia miesięczna", "Miesiące zapasu",
        "Status prognozy", "Status produktu", "Sezonowy", "Obserwowany",
        "Data zamówienia", "Data końca zapasu",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1c1917", end_color="1c1917", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p in products:
        ws.append([
            p.sku, p.name, p.ean or "", p.manufacturer_name or "", p.stock,
            p.purchase_price, p.stock_value, p.stock_in_transit, p.cbm_per_unit, p.lead_time_days,
            p.sales_1m, p.sales_2m, p.sales_3m, p.sales_4m,
            p.sales_yoy_30d, p.sales_yoy_next_30d,
            p.avg_monthly_weighted, p.months_of_stock,
            p.status, p.product_status,
            "tak" if p.seasonality_enabled else "nie",
            "tak" if p.is_favorite else "nie",
            p.order_date.isoformat(), p.empty_date.isoformat(),
        ])

    column_widths = [12, 35, 16, 15, 8, 12, 14, 10, 8, 8, 10, 10, 10, 10, 10, 10, 12, 10, 14, 14, 8, 10, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"produkty_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.put("/products/{sku:path}/favorite", response_model=ProductSummary)
async def toggle_favorite(sku: str, db: AsyncSession = Depends(get_db)):
    """Przełącza status ulubione - jeśli był true, robi false i odwrotnie."""
    existing = await db.execute(text(f"SELECT is_favorite FROM {settings.TABLE_PRODUCT_ATTRS} WHERE sku = :sku"), {"sku": sku})
    e = existing.first()
    new_val = not e.is_favorite if e else True

    await db.execute(
        text(f"""
            INSERT INTO {settings.TABLE_PRODUCT_ATTRS} (sku, is_favorite, updated_at)
            VALUES (:sku, :fav, CURRENT_TIMESTAMP)
            ON CONFLICT (sku) DO UPDATE SET is_favorite = EXCLUDED.is_favorite, updated_at = CURRENT_TIMESTAMP
        """),
        {"sku": sku, "fav": new_val}
    )
    await db.commit()
    return await get_product(db, sku)


@router.get("/favorites", response_model=List[ProductSummary])
async def list_favorites(shop: str = "", db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    """Zwraca tylko ulubione produkty."""
    products = await fetch_products(db, {"ACTIVE", "ACTIVE_NO_STOCK", "DEAD_STOCK", "INACTIVE"}, shop)
    return _mask_financials([p for p in products if p.is_favorite], user)
